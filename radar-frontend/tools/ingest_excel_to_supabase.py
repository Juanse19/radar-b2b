#!/usr/bin/env python3
"""
ingest_excel_to_supabase.py — Ingest completo de empresas desde Excel a Supabase

Fases:
  1. Empresas V2 (Aeropuertos, Cartón y Papel, Final de Línea, Solumat) — 92-93 cols
  2. Empresas V1 (Cargo LATAM, Ensambladoras Motos) — 30 cols compactas
  3. Empresas Colombia CSV — 235 filas
  4. Calificaciones históricas sintéticas (una fila por empresa con score/tier del Excel)
  5. Contactos existentes (hoja Contactos del Excel Cargo — formato Apollo)

Uso:
  pip install -r requirements.txt
  python ingest_excel_to_supabase.py [--dry-run] [--skip-errors] [--no-dedupe]
    [--only FASE]  # 1|2|3|4|5

Variables de entorno (.env de radar-frontend/):
  SUPABASE_DB_URL=postgresql://postgres:PASSWORD@db.xxx.supabase.co:5432/postgres
  DB_SCHEMA=matec_radar
"""

import os
import sys
import csv
import json
import argparse
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)
env_local = Path(__file__).parent.parent / ".env.local"
if env_local.exists():
    load_dotenv(env_local, override=True)

try:
    import psycopg2
    import psycopg2.extras
    import openpyxl
    from unidecode import unidecode
except ImportError as e:
    print(f"ERROR: Falta dependencia — {e}")
    print("Instalar con: pip install -r requirements.txt")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_URL    = os.environ.get("SUPABASE_DB_URL") or os.environ.get("DATABASE_URL")
SCHEMA    = os.environ.get("DB_SCHEMA", "matec_radar")
REPO_ROOT = Path(__file__).parent.parent.parent

DOCS_DIR  = REPO_ROOT / "docs" / "PROSPECCIÓN"

EXCEL_V2 = [
    {
        "file":        DOCS_DIR / "Linea Aeropuertos" / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",
        "sub_linea":   "aeropuertos",
        "sheet":       "Base de Datos",
        "version":     "v2",
    },
    {
        "file":        DOCS_DIR / "Linea Carton y Papel" / "BASE DE DATOS CARTON Y PAPEL.xlsx",
        "sub_linea":   "carton_corrugado",
        "sheet":       "Base de Datos",
        "version":     "v2",
    },
    {
        "file":        DOCS_DIR / "Final de Línea" / "BASE DE DATOS FINAL DE LINEA.xlsx",
        "sub_linea":   "final_linea",
        "sheet":       "Base de Datos",
        "version":     "v2",
    },
    {
        "file":        DOCS_DIR / "Línea Solumat" / "BASE DE DATOS SOLUMAT.xlsx",
        "sub_linea":   "solumat",
        "sheet":       "Base de Datos",
        "version":     "v2",
    },
]

EXCEL_V1 = [
    {
        "file":        DOCS_DIR / "Línea Cargo" / "BASE DE DATOS CARGO LATAM.xlsx",
        "sub_linea":   "cargo_uld",
        "sheet":       "Base de Datos",
        "version":     "v1",
    },
    {
        "file":        DOCS_DIR / "Ensambladora de Motos" / "BASE DE DATOS ENSAMBLADORAS MOTOS LATAM.xlsx",
        "sub_linea":   "ensambladoras_motos",
        "sheet":       "Base de Datos",
        "version":     "v1",
    },
]

COLOMBIA_CSV = DOCS_DIR / "Líneas Colombianas" / "empresas_colombia_2026.csv"

CARGO_EXCEL = DOCS_DIR / "Línea Cargo" / "BASE DE DATOS CARGO LATAM.xlsx"

# ---------------------------------------------------------------------------
# Normalización
# ---------------------------------------------------------------------------
PAIS_MAP = {
    "mexico": "MX", "méxico": "MX",
    "colombia": "CO",
    "brasil": "BR", "brazil": "BR",
    "argentina": "AR",
    "chile": "CL",
    "peru": "PE", "perú": "PE",
    "ecuador": "EC",
    "uruguay": "UY",
    "paraguay": "PY",
    "bolivia": "BO",
    "venezuela": "VE",
    "costa rica": "CR",
    "panama": "PA", "panamá": "PA",
    "guatemala": "GT",
    "el salvador": "SV",
    "honduras": "HN",
    "nicaragua": "NI",
    "republica dominicana": "DO", "república dominicana": "DO",
    "jamaica": "JM",
    "bahamas": "BS",
    "estados unidos": "US", "usa": "US", "united states": "US",
    "canada": "CA", "canadá": "CA",
    "espana": "ES", "españa": "ES",
}

TIER_MAP = {
    "tier a": "A", "tier_a": "A", "a": "A",
    "tier b": "B", "tier_b": "B", "b": "B",
    "tier b-alta": "B", "tier b alta": "B",
    "tier c": "C", "tier_c": "C", "c": "C",
    "tier d": "D", "tier_d": "D", "d": "D",
    "#n/a": "sin_calificar", "#ref!": "sin_calificar", "": "sin_calificar",
    "none": "sin_calificar", "sin calificar": "sin_calificar",
    "sin_calificar": "sin_calificar",
}

def norm_name(name: str) -> str:
    """unaccent + lower — equivalente al GENERATED ALWAYS de Postgres."""
    return unidecode(name).lower().strip()

def norm_pais(raw) -> tuple[str | None, str | None]:
    """Devuelve (pais_iso, pais_nombre)."""
    if not raw or str(raw).strip() in ("None", "", "#N/A"):
        return None, None
    text = str(raw).strip()
    iso = PAIS_MAP.get(text.lower())
    if not iso:
        # Intento con unaccent
        iso = PAIS_MAP.get(unidecode(text).lower())
    return iso or "Otro", text

def norm_tier(raw) -> str:
    if not raw:
        return "sin_calificar"
    text = str(raw).strip()
    return TIER_MAP.get(text.lower(), "sin_calificar")

def clean_score(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return round(f, 2) if 0 <= f <= 10 else None
    except (ValueError, TypeError):
        return None

def clean_val(val):
    """Convierte celdas vacías / errores a None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None", "nan", ""):
        return None
    return s

# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------
stats = {}

def init_stat(k):
    stats[k] = {"insertados": 0, "actualizados": 0, "skippeados": 0, "errores": 0}

def report():
    print("\n" + "=" * 65)
    print("  REPORTE FINAL DE INGEST")
    print("=" * 65)
    for key, s in stats.items():
        print(f"  {key:40s}  ins={s['insertados']:5d}  upd={s['actualizados']:5d}  "
              f"skip={s['skippeados']:5d}  err={s['errores']:5d}")
    print(f"\n  Total errores: {sum(s['errores'] for s in stats.values())}")

# ---------------------------------------------------------------------------
# DB
# ---------------------------------------------------------------------------
def get_conn():
    if not DB_URL:
        print("ERROR: SUPABASE_DB_URL no definida. Configura tu .env")
        sys.exit(1)
    return psycopg2.connect(DB_URL, options=f"-c search_path={SCHEMA},public")

def get_sublinea_ids(cur) -> dict:
    cur.execute(f"SELECT codigo, id FROM {SCHEMA}.sub_lineas_negocio")
    return dict(cur.fetchall())

def disable_triggers(cur):
    """Deshabilitar triggers de cache durante ingest masivo."""
    cur.execute(f"ALTER TABLE {SCHEMA}.empresas DISABLE TRIGGER trg_empresas_updated_at;")
    cur.execute(f"ALTER TABLE {SCHEMA}.calificaciones DISABLE TRIGGER trg_calif_sync;")

def enable_triggers(cur):
    cur.execute(f"ALTER TABLE {SCHEMA}.empresas ENABLE TRIGGER trg_empresas_updated_at;")
    cur.execute(f"ALTER TABLE {SCHEMA}.calificaciones ENABLE TRIGGER trg_calif_sync;")

# ---------------------------------------------------------------------------
# Upsert empresa (core)
# ---------------------------------------------------------------------------
def upsert_empresa(cur, empresa: dict, dry_run: bool, no_dedupe: bool) -> int | None:
    """
    Inserta o actualiza empresa. Devuelve empresa_id o None si dry_run.
    Dedupe por company_name_norm (con UNIQUE INDEX donde owner_id IS NULL).
    """
    name = empresa.get("company_name", "").strip()
    if not name:
        return None

    if dry_run:
        print(f"  [DRY-RUN] empresa: {name} | pais={empresa.get('pais')} | tier={empresa.get('tier_actual')}")
        return None

    cols = [
        "company_name", "pais", "pais_nombre", "estado_region", "ciudad",
        "industria_cliente", "grupo_empresarial", "marca",
        "company_domain", "company_url",
        "tier_actual", "score_total_ultimo",
        "prioridad", "pipeline",
        "meta", "keywords",
        "source_file", "source_sheet", "imported_at",
    ]
    vals = [
        name,
        empresa.get("pais"),
        empresa.get("pais_nombre"),
        empresa.get("estado_region"),
        empresa.get("ciudad"),
        empresa.get("industria_cliente"),
        empresa.get("grupo_empresarial"),
        empresa.get("marca"),
        empresa.get("company_domain"),
        empresa.get("company_url"),
        empresa.get("tier_actual", "sin_calificar"),
        empresa.get("score_total_ultimo"),
        empresa.get("prioridad", "media"),
        empresa.get("pipeline", "no_iniciado"),
        json.dumps(empresa.get("meta", {}), ensure_ascii=False),
        empresa.get("keywords"),
        empresa.get("source_file"),
        empresa.get("source_sheet"),
        datetime.utcnow().isoformat(),
    ]

    conflict_action = """
        DO UPDATE SET
          pais              = EXCLUDED.pais,
          pais_nombre       = EXCLUDED.pais_nombre,
          industria_cliente = EXCLUDED.industria_cliente,
          meta              = EXCLUDED.meta,
          source_file       = EXCLUDED.source_file,
          imported_at       = EXCLUDED.imported_at,
          updated_at        = NOW()
    """ if not no_dedupe else "DO NOTHING"

    sql = f"""
        INSERT INTO {SCHEMA}.empresas ({', '.join(cols)})
        VALUES ({', '.join(['%s'] * len(vals))})
        ON CONFLICT (company_name_norm) WHERE owner_id IS NULL
        {conflict_action}
        RETURNING id
    """
    cur.execute(sql, vals)
    row = cur.fetchone()
    return row[0] if row else None

def upsert_pivot(cur, empresa_id: int, sub_linea_id: int, es_principal: bool, dry_run: bool):
    if dry_run or not empresa_id:
        return
    cur.execute(f"""
        INSERT INTO {SCHEMA}.empresa_sub_lineas (empresa_id, sub_linea_id, es_principal)
        VALUES (%s, %s, %s)
        ON CONFLICT (empresa_id, sub_linea_id) DO NOTHING
    """, (empresa_id, sub_linea_id, es_principal))

def upsert_terminal(cur, empresa_id: int, iata: str, nombre: str | None,
                    pais_iso: str | None, dry_run: bool):
    if dry_run or not empresa_id or not iata or len(iata.strip()) != 3:
        return
    cur.execute(f"""
        INSERT INTO {SCHEMA}.empresa_terminales (empresa_id, iata_code, nombre, pais)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (empresa_id, iata_code) DO NOTHING
    """, (empresa_id, iata.strip().upper(), nombre, pais_iso))

# ---------------------------------------------------------------------------
# FASE 1 — Empresas V2
# ---------------------------------------------------------------------------
def load_sheet_rows(path: Path, sheet_name: str):
    """Carga una hoja de Excel, devuelve (header_list, rows)."""
    if not path.exists():
        return None, None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Intentar coincidir el nombre de hoja
    target = None
    for name in wb.sheetnames:
        if sheet_name.lower() in name.lower() or name.lower() in sheet_name.lower():
            target = name
            break
    if not target:
        target = wb.sheetnames[0]  # fallback primera hoja

    ws = wb[target]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    header = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(all_rows[0])]
    return header, all_rows[1:]

def parse_v2_empresa(row, header, sub_linea_code: str, source_file: str) -> dict:
    """
    Mapea una fila del Excel V2 a dict de empresa.
    Los campos no mapeados van a meta JSONB.
    """
    col = {h: i for i, h in enumerate(header)}

    def g(key_candidates, row=row, col=col):
        """get first matching column value."""
        for k in key_candidates:
            for col_name, idx in col.items():
                if k.lower() in col_name.lower():
                    v = clean_val(row[idx])
                    if v:
                        return v
        return None

    pais_raw = g(["Pais", "País", "Country"])
    pais_iso, pais_nombre = norm_pais(pais_raw)

    tier_raw   = g(["Tier", "TIER", "Clasificación"])
    score_raw  = g(["Score", "SCORE", "Puntaje", "Score_Total"])

    # Core fields
    empresa = {
        "company_name":      g(["Empresa", "Company", "Nombre_Empresa", "Nombre Empresa"]) or "",
        "pais":              pais_iso,
        "pais_nombre":       pais_nombre,
        "estado_region":     g(["Estado", "Region", "Región", "Departamento"]),
        "ciudad":            g(["Ciudad", "City"]),
        "industria_cliente": g(["Industria", "Sector", "Giro"]),
        "grupo_empresarial": g(["Grupo", "Grupo Empresarial", "Holding"]),
        "company_domain":    g(["Dominio", "Domain", "Web", "URL"]),
        "company_url":       g(["URL", "Website", "Sitio Web"]),
        "tier_actual":       norm_tier(tier_raw),
        "score_total_ultimo": clean_score(score_raw),
        "source_file":       source_file,
        "source_sheet":      "Base de Datos",
    }

    # Meta: todos los campos V2 que no se mapearon arriba
    META_INCLUDE_KEYS = [
        "Dolor_Principal", "Propuesta_Valor", "Decisor_Economico",
        "Canal_Prospeccion", "Comentarios", "Observaciones",
        "Año_Objetivo", "Ticket_Estimado", "Referente",
        "Multiplanta", "Recurrencia", "Impacto", "Prioridad_Comercial",
        "Ventana_Compra", "Presencia_Internacional", "Num_Empleados",
        "Terminales", "Aeropuertos_Operados", "Tipo_Operacion",
    ]

    meta = {}
    for col_name, idx in col.items():
        for mk in META_INCLUDE_KEYS:
            if mk.lower().replace("_", " ") in col_name.lower() or mk.lower() in col_name.lower():
                v = clean_val(row[idx])
                if v:
                    meta[col_name] = v
                break

    meta["sub_linea_codigo"]  = sub_linea_code
    meta["meta_schema"]       = "v2_amplio"

    empresa["meta"] = meta
    return empresa

def parse_v1_empresa(row, header, sub_linea_code: str, source_file: str) -> dict:
    """Mapeo compacto para los 30 campos del Excel V1."""
    col = {h.strip(): i for i, h in enumerate(header)}

    def g(*keys):
        for k in keys:
            for col_name, idx in col.items():
                if k.lower() in col_name.lower():
                    v = clean_val(row[idx])
                    if v:
                        return v
        return None

    pais_raw = g("Pais", "País", "Country")
    pais_iso, pais_nombre = norm_pais(pais_raw)

    tier_raw  = g("Tier", "TIER")
    score_raw = g("Score", "Puntaje")

    empresa = {
        "company_name":      g("Empresa", "Company", "Nombre") or "",
        "marca":             g("Marca", "Brand"),
        "pais":              pais_iso,
        "pais_nombre":       pais_nombre,
        "estado_region":     g("Estado", "Region", "Departamento"),
        "ciudad":            g("Ciudad", "City"),
        "industria_cliente": g("Industria", "Sector", "Giro"),
        "grupo_empresarial": g("Grupo", "Holding"),
        "company_domain":    g("Dominio", "Web", "URL"),
        "tier_actual":       norm_tier(tier_raw),
        "score_total_ultimo": clean_score(score_raw),
        "source_file":       source_file,
        "source_sheet":      "Base de Datos",
        "meta": {
            "sub_linea_codigo": sub_linea_code,
            "meta_schema": "v1_compacto",
            "tipo_operacion": g("Tipo_Operacion", "Tipo Operacion", "Tipo"),
        }
    }
    return empresa

# ---------------------------------------------------------------------------
# FASE 1 + 2 shared runner
# ---------------------------------------------------------------------------
def run_fase_empresas(cur, excel_entries: list, sublinea_ids: dict,
                       parse_fn, dry_run: bool, no_dedupe: bool,
                       skip_errors: bool, fase_label: str):
    init_stat(fase_label)
    key = fase_label

    for entry in excel_entries:
        path: Path  = entry["file"]
        sub_code    = entry["sub_linea"]
        sheet       = entry["sheet"]
        sub_id      = sublinea_ids.get(sub_code)

        if not path.exists():
            print(f"  SKIP {path.name} — archivo no encontrado")
            stats[key]["skippeados"] += 1
            continue
        if not sub_id:
            print(f"  SKIP {path.name} — sub_linea '{sub_code}' no en DB")
            continue

        print(f"\n[{fase_label}] {path.name} → {sub_code}")
        header, rows = load_sheet_rows(path, sheet)
        if header is None:
            print(f"  SKIP — no se pudo leer")
            continue

        for row in rows:
            if not any(row):
                continue
            try:
                empresa = parse_fn(row, header, sub_code, path.name)
                if not empresa.get("company_name"):
                    stats[key]["skippeados"] += 1
                    continue

                emp_id = upsert_empresa(cur, empresa, dry_run, no_dedupe)

                if emp_id is not None:
                    upsert_pivot(cur, emp_id, sub_id, True, dry_run)
                    stats[key]["insertados"] += 1

                    # Terminales (solo aeropuertos)
                    if sub_code == "aeropuertos":
                        terminales_raw = empresa.get("meta", {}).get("Terminales") or \
                                         empresa.get("meta", {}).get("Aeropuertos_Operados")
                        if terminales_raw:
                            for part in str(terminales_raw).split(","):
                                part = part.strip()
                                if len(part) == 3 and part.isalpha():
                                    upsert_terminal(cur, emp_id, part, None,
                                                    empresa.get("pais"), dry_run)
                elif not dry_run:
                    # Empresa ya existía — agregar pivot sin es_principal
                    cur.execute(f"""
                        SELECT id FROM {SCHEMA}.empresas
                        WHERE company_name_norm = lower(unaccent(%s))
                          AND owner_id IS NULL
                    """, (empresa["company_name"],))
                    row_db = cur.fetchone()
                    if row_db:
                        upsert_pivot(cur, row_db[0], sub_id, False, dry_run)
                    stats[key]["actualizados"] += 1

            except Exception as e:
                stats[key]["errores"] += 1
                if not skip_errors:
                    raise
                print(f"  ERROR row: {e}")

        if not dry_run:
            cur.connection.commit()
            print(f"  Commit OK — sub_linea {sub_code}")

# ---------------------------------------------------------------------------
# FASE 3 — Colombia CSV
# ---------------------------------------------------------------------------
CSV_SOURCE_TO_SUBLINEA = {
    "ALIMENTICIO": "final_linea",
    "BONATO":      "carton_corrugado",
    "CARTON":      "carton_corrugado",
    "INTRALOGIS":  "final_linea",
    "MOTOS":       "ensambladoras_motos",
    "BHS":         "aeropuertos",
}

def run_fase3_colombia(cur, sublinea_ids: dict, dry_run: bool,
                       no_dedupe: bool, skip_errors: bool):
    init_stat("colombia_csv")
    key = "colombia_csv"

    if not COLOMBIA_CSV.exists():
        print(f"\n  SKIP Fase 3 — {COLOMBIA_CSV} no encontrado")
        stats[key]["skippeados"] += 1
        return

    print(f"\n[Fase 3 Colombia] {COLOMBIA_CSV.name}")
    with open(COLOMBIA_CSV, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row_raw in reader:
            try:
                source_raw  = (row_raw.get("source") or row_raw.get("fuente") or "").upper()
                sub_code    = None
                for k, v in CSV_SOURCE_TO_SUBLINEA.items():
                    if k in source_raw:
                        sub_code = v
                        break
                sub_code = sub_code or "final_linea"
                sub_id   = sublinea_ids.get(sub_code)

                pais_raw  = row_raw.get("pais") or row_raw.get("country") or "CO"
                pais_iso, pais_nombre = norm_pais(pais_raw)

                tier_raw  = row_raw.get("tier") or row_raw.get("Tier") or ""
                score_raw = row_raw.get("score") or row_raw.get("Score")

                empresa = {
                    "company_name":       (row_raw.get("empresa") or row_raw.get("company") or "").strip(),
                    "pais":               pais_iso or "CO",
                    "pais_nombre":        pais_nombre or "Colombia",
                    "ciudad":             clean_val(row_raw.get("ciudad") or row_raw.get("city")),
                    "industria_cliente":  clean_val(row_raw.get("industria") or row_raw.get("sector")),
                    "company_domain":     clean_val(row_raw.get("dominio") or row_raw.get("web")),
                    "tier_actual":        norm_tier(tier_raw),
                    "score_total_ultimo": clean_score(score_raw),
                    "source_file":        COLOMBIA_CSV.name,
                    "source_sheet":       "csv",
                    "meta": {
                        "sub_linea_codigo": sub_code,
                        "fuente_colombia":  True,
                        "source_csv":       source_raw,
                    }
                }

                if not empresa["company_name"]:
                    stats[key]["skippeados"] += 1
                    continue

                emp_id = upsert_empresa(cur, empresa, dry_run, no_dedupe)
                if emp_id and sub_id:
                    upsert_pivot(cur, emp_id, sub_id, True, dry_run)
                    stats[key]["insertados"] += 1
                else:
                    stats[key]["actualizados"] += 1

            except Exception as e:
                stats[key]["errores"] += 1
                if not skip_errors:
                    raise
                print(f"  ERROR CSV row: {e}")

    if not dry_run:
        cur.connection.commit()

# ---------------------------------------------------------------------------
# FASE 4 — Calificaciones históricas sintéticas
# ---------------------------------------------------------------------------
def run_fase4_calificaciones_historicas(cur, dry_run: bool, skip_errors: bool):
    """
    Para cada empresa con score_total_ultimo y tier_actual != 'sin_calificar',
    inserta 1 fila sintética en calificaciones (si no existe ya una con prompt_version='import_excel_v1').
    """
    init_stat("calificaciones_historicas")
    key = "calificaciones_historicas"

    if dry_run:
        print("\n[Fase 4] DRY-RUN — se insertarían calificaciones históricas")
        return

    print("\n[Fase 4] Insertando calificaciones históricas sintéticas...")
    cur.execute(f"""
        SELECT id, tier_actual, score_total_ultimo
        FROM {SCHEMA}.empresas
        WHERE tier_actual != 'sin_calificar'
          AND score_total_ultimo IS NOT NULL
          AND id NOT IN (
            SELECT DISTINCT empresa_id FROM {SCHEMA}.calificaciones
            WHERE prompt_version = 'import_excel_v1'
          )
    """)
    empresas = cur.fetchall()
    print(f"  {len(empresas)} empresas necesitan calificación histórica")

    for emp_id, tier, score in empresas:
        try:
            cur.execute(f"""
                INSERT INTO {SCHEMA}.calificaciones
                  (empresa_id, score_total, tier_calculado, razonamiento_agente,
                   prompt_version, modelo_llm)
                VALUES (%s, %s, %s, %s, 'import_excel_v1', 'import')
                ON CONFLICT DO NOTHING
            """, (emp_id, score, tier,
                  "Calificación importada desde base de datos Excel inicial (pre-Supabase)"))
            stats[key]["insertados"] += 1
        except Exception as e:
            stats[key]["errores"] += 1
            if not skip_errors:
                raise
            print(f"  ERROR empresa {emp_id}: {e}")

    # Re-enable triggers y hacer refresh del cache
    enable_triggers(cur)

    # Sync cache manualmente: ejecutar UPDATE masivo para los que no lo tienen
    cur.execute(f"""
        UPDATE {SCHEMA}.empresas e
        SET ultima_calificacion_id = c.id,
            ultima_calificacion_at = c.created_at,
            score_total_ultimo     = c.score_total,
            tier_actual            = c.tier_calculado,
            updated_at             = NOW()
        FROM (
            SELECT DISTINCT ON (empresa_id)
                   id, empresa_id, score_total, tier_calculado, created_at
            FROM {SCHEMA}.calificaciones
            ORDER BY empresa_id, created_at DESC
        ) c
        WHERE e.id = c.empresa_id
    """)

    cur.connection.commit()
    print(f"  Cache empresas actualizado")

# ---------------------------------------------------------------------------
# FASE 5 — Contactos existentes (Cargo Excel)
# ---------------------------------------------------------------------------
def run_fase5_contactos_cargo(cur, sublinea_ids: dict, dry_run: bool, skip_errors: bool):
    """
    Lee la hoja 'Contactos' del Excel de Cargo LATAM (formato Apollo export).
    """
    init_stat("contactos_cargo")
    key = "contactos_cargo"

    if not CARGO_EXCEL.exists():
        print(f"\n  SKIP Fase 5 — {CARGO_EXCEL} no encontrado")
        return

    wb = openpyxl.load_workbook(CARGO_EXCEL, read_only=True, data_only=True)
    contact_sheet = None
    for name in wb.sheetnames:
        if "contacto" in name.lower():
            contact_sheet = name
            break
    if not contact_sheet:
        print(f"\n  SKIP Fase 5 — hoja 'Contactos' no encontrada en {CARGO_EXCEL.name}")
        return

    print(f"\n[Fase 5] Contactos desde {CARGO_EXCEL.name} → hoja {contact_sheet}")
    ws = wb[contact_sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return

    header = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(rows[0])]
    col = {h: i for i, h in enumerate(header)}

    def g(*keys):
        for k in keys:
            for col_name, idx in col.items():
                if k.lower() in col_name.lower():
                    return clean_val(row[idx])
        return None

    for row in rows[1:]:
        if not any(row):
            continue

        try:
            empresa_name = g("Empresa", "Company") or ""
            if not empresa_name:
                stats[key]["skippeados"] += 1
                continue

            # Buscar empresa_id
            if not dry_run:
                cur.execute(f"""
                    SELECT id FROM {SCHEMA}.empresas
                    WHERE company_name_norm = lower(unaccent(%s))
                      AND owner_id IS NULL
                    LIMIT 1
                """, (empresa_name,))
                emp_row = cur.fetchone()
                if not emp_row:
                    stats[key]["skippeados"] += 1
                    continue
                emp_id = emp_row[0]
            else:
                emp_id = 0

            pais_raw = g("Country", "Pais", "País")
            pais_iso, _ = norm_pais(pais_raw)

            first_name = g("First Name", "Nombre", "first_name")
            last_name  = g("Last Name",  "Apellido", "last_name")
            email      = g("Email", "Email_Verificado", "Correo")
            apollo_id  = g("Apollo ID", "apollo_id", "Person_ID", "Persona_ID")

            if dry_run:
                print(f"  [DRY-RUN] contacto: {first_name} {last_name} | {empresa_name}")
                stats[key]["insertados"] += 1
                continue

            cur.execute(f"""
                INSERT INTO {SCHEMA}.contactos
                  (empresa_id, first_name, last_name, title, email, email_status,
                   phone_work_direct, phone_mobile, linkedin_url,
                   city, state, country, apollo_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (apollo_id) DO UPDATE SET
                  email      = EXCLUDED.email,
                  updated_at = NOW()
                RETURNING (xmax = 0)
            """, (
                emp_id,
                first_name, last_name,
                g("Title", "Cargo", "Titulo"),
                email,
                g("Email Status", "Estado_Email"),
                g("Work Phone", "Tel_Empresa", "Corporate Phone"),
                g("Mobile Phone", "Tel_Movil"),
                g("LinkedIn", "LinkedIn URL", "linkedin_url"),
                g("City", "Ciudad"),
                g("State", "Estado"),
                pais_iso,
                apollo_id,
            ))
            is_insert = cur.fetchone()
            if is_insert and is_insert[0]:
                stats[key]["insertados"] += 1
            else:
                stats[key]["actualizados"] += 1

        except Exception as e:
            stats[key]["errores"] += 1
            if not skip_errors:
                raise
            print(f"  ERROR contacto: {e}")

    if not dry_run:
        cur.connection.commit()

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Ingest completo de Excel → Supabase")
    parser.add_argument("--dry-run",     action="store_true")
    parser.add_argument("--skip-errors", action="store_true")
    parser.add_argument("--no-dedupe",   action="store_true",
                        help="Usar INSERT DO NOTHING en vez de DO UPDATE (debug)")
    parser.add_argument("--only", choices=["1","2","3","4","5"],
                        help="Ejecutar solo la fase indicada")
    args = parser.parse_args()

    print(f"{'[DRY-RUN] ' if args.dry_run else ''}Ingest Excel → Supabase ({SCHEMA})")
    conn = None if args.dry_run else get_conn()
    cur  = None if args.dry_run else conn.cursor()

    if not args.dry_run:
        disable_triggers(cur)
        sublinea_ids = get_sublinea_ids(cur)
        conn.commit()
    else:
        sublinea_ids = {
            "aeropuertos": 1, "cargo_uld": 2, "carton_corrugado": 3,
            "final_linea": 4, "ensambladoras_motos": 5, "solumat": 6,
        }

    try:
        only = args.only

        if not only or only == "1":
            print("\n=== FASE 1: Empresas V2 ===")
            run_fase_empresas(cur, EXCEL_V2, sublinea_ids, parse_v2_empresa,
                              args.dry_run, args.no_dedupe, args.skip_errors,
                              "empresas_v2")

        if not only or only == "2":
            print("\n=== FASE 2: Empresas V1 ===")
            run_fase_empresas(cur, EXCEL_V1, sublinea_ids, parse_v1_empresa,
                              args.dry_run, args.no_dedupe, args.skip_errors,
                              "empresas_v1")

        if not only or only == "3":
            print("\n=== FASE 3: Empresas Colombia CSV ===")
            run_fase3_colombia(cur, sublinea_ids, args.dry_run,
                               args.no_dedupe, args.skip_errors)

        if not only or only == "4":
            print("\n=== FASE 4: Calificaciones históricas ===")
            run_fase4_calificaciones_historicas(cur, args.dry_run, args.skip_errors)
        elif not args.dry_run:
            enable_triggers(cur)
            conn.commit()

        if not only or only == "5":
            print("\n=== FASE 5: Contactos Cargo ===")
            run_fase5_contactos_cargo(cur, sublinea_ids, args.dry_run, args.skip_errors)

    except Exception as e:
        if conn:
            conn.rollback()
        if cur:
            enable_triggers(cur)
            conn.commit()
        print(f"\nFATAL: {e}")
        raise
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    report()
    total_err = sum(s['errores'] for s in stats.values())
    sys.exit(1 if total_err > 0 else 0)

if __name__ == "__main__":
    main()
