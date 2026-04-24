#!/usr/bin/env python3
"""
seed_catalogos.py — Fase 1 del ingest: catálogos de gran volumen
  - sectores (~3559 filas desde hoja 'Datos' del Excel Aeropuertos)
  - job_titles_por_linea (desde docs/docs_contactos/JobTitles_PorLineaNegocio.xlsx)
  - palabras_clave_por_linea (desde hoja 'Palabras Clave' de cada Excel V2)

Uso:
  pip install -r requirements.txt
  python seed_catalogos.py [--dry-run] [--skip-errors]

Variables de entorno (.env):
  SUPABASE_DB_URL=postgresql://postgres:PASSWORD@db.xxx.supabase.co:5432/postgres
  DB_SCHEMA=matec_radar
"""

import os
import sys
import argparse
import json
from pathlib import Path
from dotenv import load_dotenv

# Cargar .env desde radar-frontend/
env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)
env_local = Path(__file__).parent.parent / ".env.local"
if env_local.exists():
    load_dotenv(env_local, override=True)

try:
    import psycopg2
    import psycopg2.extras
    import openpyxl
except ImportError as e:
    print(f"ERROR: Falta dependencia — {e}")
    print("Instalar con: pip install -r requirements.txt")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DB_URL    = os.environ.get("SUPABASE_DB_URL") or os.environ.get("DATABASE_URL")
SCHEMA    = os.environ.get("DB_SCHEMA", "matec_radar")
REPO_ROOT = Path(__file__).parent.parent.parent  # clients/

DOCS_DIR  = REPO_ROOT / "docs" / "PROSPECCIÓN"
JOBS_FILE = REPO_ROOT / "docs" / "docs_contactos" / "JobTitles_PorLineaNegocio.xlsx"

EXCEL_V2_FILES = {
    "aeropuertos":  DOCS_DIR / "Linea Aeropuertos" / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",
    "carton_corrugado": DOCS_DIR / "Linea Carton y Papel" / "BASE DE DATOS CARTON Y PAPEL.xlsx",
    "final_linea":  DOCS_DIR / "Final de Línea" / "BASE DE DATOS FINAL DE LINEA.xlsx",
    "solumat":      DOCS_DIR / "Línea Solumat" / "BASE DE DATOS SOLUMAT.xlsx",
}

# Mapeo hoja → sub_linea codigo (para job titles desde Excel con varias hojas)
JT_SHEET_TO_SUBLINEA = {
    "Aeropuertos":          "aeropuertos",
    "Cargo":                "cargo_uld",
    "Cargo ULD":            "cargo_uld",
    "Carton":               "carton_corrugado",
    "Cartón":               "carton_corrugado",
    "Carton y Papel":       "carton_corrugado",
    "Cartón y Papel":       "carton_corrugado",
    "Final de Linea":       "final_linea",
    "Final de Línea":       "final_linea",
    "Motos":                "ensambladoras_motos",
    "Ensambladoras Motos":  "ensambladoras_motos",
    "Solumat":              "solumat",
    "Intralogistica":       "final_linea",  # fallback
    "BHS":                  "aeropuertos",  # fallback
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
stats = {}

def init_stat(key):
    stats[key] = {"insertados": 0, "actualizados": 0, "skippeados": 0, "errores": 0}

def report():
    print("\n{'=' * 60}")
    print("  REPORTE DE SEED")
    print("{'=' * 60}")
    for key, s in stats.items():
        print(f"  {key:35s}  ins={s['insertados']:4d}  upd={s['actualizados']:4d}  "
              f"skip={s['skippeados']:4d}  err={s['errores']:4d}")
    total_err = sum(s['errores'] for s in stats.values())
    print(f"\n  Total errores: {total_err}")

def get_conn():
    if not DB_URL:
        print("ERROR: SUPABASE_DB_URL no definida. Agrega al .env de radar-frontend.")
        sys.exit(1)
    return psycopg2.connect(DB_URL, options=f"-c search_path={SCHEMA},public")

def get_sublinea_ids(cur):
    cur.execute(f"SELECT codigo, id FROM {SCHEMA}.sub_lineas_negocio")
    return dict(cur.fetchall())

# ---------------------------------------------------------------------------
# Fase 1: Sectores
# ---------------------------------------------------------------------------
def seed_sectores(cur, dry_run: bool, skip_errors: bool):
    """
    Lee la hoja 'Datos' del Excel de Aeropuertos.
    Espera columnas: Código_Sector (o ID_Sector), Nombre_Sector, Nivel (opcional)
    """
    init_stat("sectores")
    key = "sectores"

    excel_path = EXCEL_V2_FILES["aeropuertos"]
    if not excel_path.exists():
        print(f"  SKIP sectores — archivo no encontrado: {excel_path}")
        stats[key]["skippeados"] += 1
        return

    print(f"\n[sectores] Leyendo {excel_path.name}...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Encontrar hoja Datos
    sheet_name = None
    for name in wb.sheetnames:
        if "dato" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        print(f"  SKIP sectores — hoja 'Datos' no encontrada en {excel_path.name}")
        stats[key]["skippeados"] += 1
        return

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    # Detectar headers
    header = [str(c).strip() if c else "" for c in rows[0]]
    col = {h: i for i, h in enumerate(header)}

    # Intentar detectar columnas de sector
    nombre_col = next(
        (c for c in col if any(k in c.lower() for k in ["sector", "industria", "nombre_sector"])),
        None
    )
    codigo_col = next(
        (c for c in col if any(k in c.lower() for k in ["codigo", "código", "id_sector", "cod"])),
        None
    )

    if not nombre_col:
        print(f"  SKIP sectores — no se encontró columna de nombre en headers: {header[:10]}")
        stats[key]["skippeados"] += 1
        return

    seen = set()
    for row in rows[1:]:
        nombre = str(row[col[nombre_col]]).strip() if nombre_col and row[col[nombre_col]] else None
        codigo = str(row[col[codigo_col]]).strip() if codigo_col and row[col[codigo_col]] else None

        if not nombre or nombre in ("None", "#N/A", "#REF!", ""):
            continue
        if nombre in seen:
            stats[key]["skippeados"] += 1
            continue
        seen.add(nombre)

        if dry_run:
            print(f"  [DRY-RUN] sector: {codigo or '?'} | {nombre}")
            stats[key]["insertados"] += 1
            continue

        try:
            cur.execute(f"""
                INSERT INTO {SCHEMA}.sectores (codigo, nombre)
                VALUES (%s, %s)
                ON CONFLICT (codigo) DO UPDATE SET nombre = EXCLUDED.nombre
                RETURNING (xmax = 0)
            """, (codigo, nombre))
            is_insert = cur.fetchone()[0]
            if is_insert:
                stats[key]["insertados"] += 1
            else:
                stats[key]["actualizados"] += 1
        except Exception as e:
            stats[key]["errores"] += 1
            if not skip_errors:
                raise
            print(f"  ERROR sector '{nombre}': {e}")

    print(f"  sectores procesados: {len(seen)}")

# ---------------------------------------------------------------------------
# Fase 2: Job Titles
# ---------------------------------------------------------------------------
def seed_job_titles(cur, dry_run: bool, skip_errors: bool):
    """
    Lee el archivo JobTitles_PorLineaNegocio.xlsx.
    Cada hoja = una línea de negocio.
    Columnas esperadas: Titulo, Nivel (1-5), Idioma (es/en), Prioridad (1-3)
    """
    init_stat("job_titles_por_linea")
    key = "job_titles_por_linea"

    if not JOBS_FILE.exists():
        print(f"  SKIP job_titles — archivo no encontrado: {JOBS_FILE}")
        stats[key]["skippeados"] += 1
        return

    sublinea_ids = get_sublinea_ids(cur)
    print(f"\n[job_titles] Leyendo {JOBS_FILE.name}...")
    wb = openpyxl.load_workbook(JOBS_FILE, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        sublinea_codigo = JT_SHEET_TO_SUBLINEA.get(sheet_name)
        if not sublinea_codigo:
            # Buscar por coincidencia parcial
            for map_key, map_val in JT_SHEET_TO_SUBLINEA.items():
                if map_key.lower() in sheet_name.lower() or sheet_name.lower() in map_key.lower():
                    sublinea_codigo = map_val
                    break
        if not sublinea_codigo:
            print(f"  SKIP hoja '{sheet_name}' — sin mapeo a sub-línea")
            continue

        sub_id = sublinea_ids.get(sublinea_codigo)
        if not sub_id:
            print(f"  SKIP hoja '{sheet_name}' — sub_linea '{sublinea_codigo}' no existe en DB")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        col = {h: i for i, h in enumerate(header)}

        titulo_col  = next((c for c in col if "titulo" in c or "title" in c), None)
        nivel_col   = next((c for c in col if "nivel" in c or "level" in c), None)
        idioma_col  = next((c for c in col if "idioma" in c or "lang" in c), None)
        prio_col    = next((c for c in col if "prior" in c), None)

        if not titulo_col:
            print(f"  SKIP hoja '{sheet_name}' — sin columna 'titulo'")
            continue

        for row in rows[1:]:
            titulo = str(row[col[titulo_col]]).strip() if row[col.get(titulo_col, 0)] else None
            if not titulo or titulo in ("None", "", "#N/A"):
                continue

            nivel    = int(row[col[nivel_col]]) if nivel_col and row[col[nivel_col]] else 3
            idioma   = str(row[col[idioma_col]]).strip()[:2] if idioma_col and row[col[idioma_col]] else "es"
            prioridad = int(row[col[prio_col]]) if prio_col and row[col[prio_col]] else 2

            # Sanitize
            nivel     = max(1, min(5, nivel))
            prioridad = max(1, min(3, prioridad))
            if idioma not in ("es", "en", "pt"):
                idioma = "es"

            if dry_run:
                print(f"  [DRY-RUN] job_title: {sublinea_codigo} | nivel={nivel} | {titulo}")
                stats[key]["insertados"] += 1
                continue

            try:
                cur.execute(f"""
                    INSERT INTO {SCHEMA}.job_titles_por_linea
                      (sub_linea_id, titulo, nivel, idioma, prioridad)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (sub_linea_id, titulo, idioma) DO UPDATE SET
                      nivel     = EXCLUDED.nivel,
                      prioridad = EXCLUDED.prioridad
                    RETURNING (xmax = 0)
                """, (sub_id, titulo, nivel, idioma, prioridad))
                is_insert = cur.fetchone()[0]
                if is_insert:
                    stats[key]["insertados"] += 1
                else:
                    stats[key]["actualizados"] += 1
            except Exception as e:
                stats[key]["errores"] += 1
                if not skip_errors:
                    raise
                print(f"  ERROR job_title '{titulo}': {e}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Seed catálogos grandes a Supabase")
    parser.add_argument("--dry-run",     action="store_true", help="Solo imprimir, no ejecutar SQL")
    parser.add_argument("--skip-errors", action="store_true", help="Continuar ante errores por fila")
    parser.add_argument("--only",        choices=["sectores", "job_titles"], help="Ejecutar solo una fase")
    args = parser.parse_args()

    print(f"{'DRY-RUN: ' if args.dry_run else ''}Conectando a Supabase...")
    conn = None if args.dry_run else get_conn()
    cur  = None if args.dry_run else conn.cursor()

    try:
        if not args.only or args.only == "sectores":
            seed_sectores(cur, args.dry_run, args.skip_errors)
            if not args.dry_run:
                conn.commit()

        if not args.only or args.only == "job_titles":
            seed_job_titles(cur, args.dry_run, args.skip_errors)
            if not args.dry_run:
                conn.commit()

    except Exception as e:
        if conn:
            conn.rollback()
        print(f"\nFATAL: {e}")
        raise
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    report()
    total_err = sum(s['errores'] for s in stats.values())
    sys.exit(1 if total_err > 0 else 0)

if __name__ == "__main__":
    main()
