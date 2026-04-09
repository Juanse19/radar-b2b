#!/usr/bin/env python3
"""
seed_via_rest.py
Seeds sectores and job_titles_por_linea via Supabase PostgREST REST API.
No direct PostgreSQL connection needed — uses SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY.

Usage:
  python tools/seed_via_rest.py [--dry-run] [--only sectores|job_titles]
"""

import os
import sys
import json
import time
import argparse
from pathlib import Path
from dotenv import load_dotenv

# Load .env.local from radar-frontend/
env_local = Path(__file__).parent.parent / ".env.local"
if env_local.exists():
    load_dotenv(env_local, override=True)

try:
    import openpyxl
    import urllib.request
    import urllib.error
except ImportError as e:
    print(f"ERROR: {e}\npip install openpyxl python-dotenv")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUPA_URL   = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPA_KEY   = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SCHEMA     = os.environ.get("SUPABASE_DB_SCHEMA", "matec_radar")
REPO_ROOT  = Path(__file__).parent.parent.parent  # clients/
DOCS_DIR   = REPO_ROOT / "docs" / "PROSPECCIÓN"
JOBS_FILE  = REPO_ROOT / "docs" / "docs_contactos" / "JobTitles_PorLineaNegocio.xlsx"

if not SUPA_URL or not SUPA_KEY:
    print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set in .env.local")
    sys.exit(1)

EXCEL_V2_FILES = {
    "aeropuertos":      DOCS_DIR / "Linea Aeropuertos"    / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",
    "carton_corrugado": DOCS_DIR / "Linea Carton y Papel"  / "BASE DE DATOS CARTON Y PAPEL.xlsx",
    "final_linea":      DOCS_DIR / "Final de Línea"        / "BASE DE DATOS FINAL DE LINEA.xlsx",
    "solumat":          DOCS_DIR / "Línea Solumat"         / "BASE DE DATOS SOLUMAT.xlsx",
}

JT_SHEET_TO_SUBLINEA = {
    "Aeropuertos":         "aeropuertos",
    "Cargo":               "cargo_uld",
    "Cargo ULD":           "cargo_uld",
    "Carton":              "carton_corrugado",
    "Cartón":              "carton_corrugado",
    "Carton y Papel":      "carton_corrugado",
    "Cartón y Papel":      "carton_corrugado",
    "Final de Linea":      "final_linea",
    "Final de Línea":      "final_linea",
    "Motos":               "ensambladoras_motos",
    "Ensambladoras Motos": "ensambladoras_motos",
    "Solumat":             "solumat",
    "Intralogistica":      "final_linea",
    "BHS":                 "aeropuertos",
}

# ---------------------------------------------------------------------------
# REST helpers
# ---------------------------------------------------------------------------
HEADERS = {
    "Authorization":  f"Bearer {SUPA_KEY}",
    "apikey":          SUPA_KEY,
    "Content-Type":   "application/json",
    "Accept":         "application/json",
    "Content-Profile": SCHEMA,
    "Accept-Profile":  SCHEMA,
    "Prefer":          "return=minimal,resolution=merge-duplicates",
    "User-Agent":      "Mozilla/5.0 (compatible; MatecRadar/1.0)",
}

def rest_post(table: str, rows: list, dry_run: bool) -> dict:
    """Bulk insert/upsert via PostgREST."""
    if dry_run:
        return {"inserted": len(rows), "dry_run": True}
    if not rows:
        return {"inserted": 0}

    BATCH = 500
    total = {"inserted": 0, "errors": 0}
    for i in range(0, len(rows), BATCH):
        batch = rows[i:i + BATCH]
        data  = json.dumps(batch).encode("utf-8")
        req   = urllib.request.Request(
            f"{SUPA_URL}/rest/v1/{table}",
            data=data,
            headers={**HEADERS, "Content-Length": str(len(data))},
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                total["inserted"] += len(batch)
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="replace")
            print(f"  HTTP {e.code} on batch {i//BATCH}: {body[:200]}")
            total["errors"] += len(batch)
            time.sleep(1)
    return total

def rest_get(table: str, select: str = "*", params: str = "") -> list:
    """GET from PostgREST."""
    qs = f"?select={select}" + (f"&{params}" if params else "")
    req = urllib.request.Request(
        f"{SUPA_URL}/rest/v1/{table}{qs}",
        headers={**HEADERS, "Content-Profile": SCHEMA, "Accept-Profile": SCHEMA},
        method="GET",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))

# ---------------------------------------------------------------------------
# Fase 1: Sectores
# ---------------------------------------------------------------------------
def seed_sectores(dry_run: bool):
    excel_path = EXCEL_V2_FILES["aeropuertos"]
    if not excel_path.exists():
        print(f"  SKIP sectores — archivo no encontrado: {excel_path}")
        return

    print(f"\n[sectores] Leyendo {excel_path.name}...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Find exact 'Datos' sheet (avoid matching 'Base de Datos')
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "datos":
            sheet_name = name
            break
    if not sheet_name:
        # Fallback: any sheet with 'dato' that is NOT 'base de datos'
        for name in wb.sheetnames:
            if "dato" in name.lower() and "base" not in name.lower():
                sheet_name = name
                break

    if not sheet_name:
        print(f"  SKIP sectores — hoja 'Datos' no encontrada. Sheets: {wb.sheetnames}")
        return

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    print(f"  Hoja '{sheet_name}': {len(rows)-1} filas")

    seen = set()
    records = []
    for row in rows[1:]:
        nombre = str(row[0]).strip() if row and row[0] else None
        if not nombre or nombre in ("None", "#N/A", "#REF!", ""):
            continue
        if nombre in seen:
            continue
        seen.add(nombre)
        records.append({"nombre": nombre, "nivel": 1, "activo": True})

    print(f"  Sectores únicos: {len(records)}")
    if dry_run:
        print(f"  [DRY-RUN] Se insertarían {len(records)} sectores")
        return

    result = rest_post("sectores", records, dry_run=False)
    print(f"  ✓ Insertados: {result['inserted']} | Errores: {result.get('errors', 0)}")

# ---------------------------------------------------------------------------
# Fase 2: Job Titles
# ---------------------------------------------------------------------------
def seed_job_titles(dry_run: bool):
    if not JOBS_FILE.exists():
        print(f"  SKIP job_titles — archivo no encontrado: {JOBS_FILE}")
        return

    print(f"\n[job_titles] Leyendo {JOBS_FILE.name}...")

    # Get sub_linea id map from DB (or fake for dry-run)
    if dry_run:
        sublinea_ids = {
            "aeropuertos": 1, "cargo_uld": 2, "carton_corrugado": 3,
            "final_linea": 4, "ensambladoras_motos": 5, "solumat": 6,
        }
    else:
        rows = rest_get("sub_lineas_negocio", select="id,codigo")
        sublinea_ids = {r["codigo"]: r["id"] for r in rows}
        print(f"  Sub-líneas en DB: {list(sublinea_ids.keys())}")

    wb = openpyxl.load_workbook(JOBS_FILE, read_only=True, data_only=True)
    total_inserted = 0

    # Normalize sheet name for matching: lowercase, remove accents, replace _ with space
    def norm_sheet(s):
        return s.lower().replace("_", " ").replace("-", " ").strip()

    for sheet_name in wb.sheetnames:
        sublinea_codigo = JT_SHEET_TO_SUBLINEA.get(sheet_name)
        if not sublinea_codigo:
            norm = norm_sheet(sheet_name)
            for map_key, map_val in JT_SHEET_TO_SUBLINEA.items():
                mk = norm_sheet(map_key)
                if mk == norm or mk in norm or norm in mk:
                    sublinea_codigo = map_val
                    break
        # Extra patterns for uppercase underscore sheet names
        if not sublinea_codigo:
            upper = sheet_name.upper().replace("_", "").replace(" ", "")
            patterns = {
                "FINALLINEA": "final_linea", "AEROPUERTOS": "aeropuertos",
                "CARGO": "cargo_uld", "CARTONPAPEL": "carton_corrugado",
                "CARTON": "carton_corrugado", "MOTOS": "ensambladoras_motos",
                "SOLUMAT": "solumat",
            }
            for p, v in patterns.items():
                if p in upper:
                    sublinea_codigo = v
                    break
        if not sublinea_codigo:
            print(f"  SKIP hoja '{sheet_name}' — sin mapeo")
            continue

        sub_id = sublinea_ids.get(sublinea_codigo)
        if not sub_id:
            print(f"  SKIP hoja '{sheet_name}' — sub_linea '{sublinea_codigo}' no en DB")
            continue

        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))
        if len(rows_data) < 2:
            continue

        header = [str(c).strip().lower() if c else "" for c in rows_data[0]]
        col = {h: i for i, h in enumerate(header)}

        # Detect column names
        titulo_col  = next((c for c in col if "titulo" in c or "title" in c or "cargo" in c), None)
        nivel_col   = next((c for c in col if "nivel" in c or "level" in c), None)
        idioma_col  = next((c for c in col if "idioma" in c or "lang" in c), None)
        prio_col    = next((c for c in col if "prioridad" in c or "priority" in c), None)

        if not titulo_col:
            # Assume first column is title
            titulo_col = header[0] if header else None
        if not titulo_col:
            print(f"  SKIP hoja '{sheet_name}' — no se detectó columna de título")
            continue

        records = []
        seen = set()
        for row in rows_data[1:]:
            titulo  = str(row[col[titulo_col]]).strip() if row[col[titulo_col]] else None
            nivel   = int(row[col[nivel_col]])   if nivel_col and row[col[nivel_col]]  else 3
            idioma  = str(row[col[idioma_col]]).strip()[:2] if idioma_col and row[col[idioma_col]] else "es"
            prio    = int(row[col[prio_col]])    if prio_col and row[col[prio_col]]    else 2

            if not titulo or titulo in ("None", "#N/A", ""):
                continue
            key = (sub_id, titulo.lower(), idioma)
            if key in seen:
                continue
            seen.add(key)
            records.append({
                "sub_linea_id": sub_id,
                "titulo":       titulo,
                "nivel":        min(max(int(nivel), 1), 5),
                "idioma":       idioma[:2],
                "prioridad":    min(max(int(prio), 1), 3),
                "activo":       True,
            })

        if dry_run:
            print(f"  [DRY-RUN] {sheet_name} -> {sublinea_codigo}: {len(records)} titulos")
            total_inserted += len(records)
            continue

        result = rest_post("job_titles_por_linea", records, dry_run=False)
        print(f"  {sheet_name} -> {sublinea_codigo}: {result['inserted']} insertados")
        total_inserted += result["inserted"]

    print(f"\n  Total job_titles: {total_inserted}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--only", choices=["sectores", "job_titles"])
    args = ap.parse_args()

    print(f"Supabase: {SUPA_URL}  schema={SCHEMA}  {'[DRY-RUN]' if args.dry_run else ''}\n")

    if not args.only or args.only == "sectores":
        seed_sectores(args.dry_run)
    if not args.only or args.only == "job_titles":
        seed_job_titles(args.dry_run)

    print("\nDone.")

if __name__ == "__main__":
    main()
