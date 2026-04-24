#!/usr/bin/env python3
"""
seed_via_pgquery.py
Seeds sectores and job_titles_por_linea using the pg/query HTTP API.
No PostgREST schema exposure required.

Usage:
  python tools/seed_via_pgquery.py [--dry-run] [--only sectores|job_titles]
"""

import os
import sys
import json
import time
import argparse
from pathlib import Path
from dotenv import load_dotenv

env_local = Path(__file__).parent.parent / ".env.local"
if env_local.exists():
    load_dotenv(env_local, override=True)

try:
    import openpyxl
    import urllib.request
    import urllib.error
except ImportError as e:
    print(f"ERROR: {e}\npip install openpyxl python-dotenv")
    sys.exit(1)

SUPA_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPA_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SCHEMA   = "matec_radar"
REPO_ROOT = Path(__file__).parent.parent.parent
DOCS_DIR  = REPO_ROOT / "docs" / "PROSPECCIÓN"
JOBS_FILE = REPO_ROOT / "docs" / "docs_contactos" / "JobTitles_PorLineaNegocio.xlsx"

EXCEL_V2_FILES = {
    "aeropuertos":      DOCS_DIR / "Linea Aeropuertos"    / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",
    "carton_corrugado": DOCS_DIR / "Linea Carton y Papel"  / "BASE DE DATOS CARTON Y PAPEL.xlsx",
    "final_linea":      DOCS_DIR / "Final de Línea"        / "BASE DE DATOS FINAL DE LINEA.xlsx",
    "solumat":          DOCS_DIR / "Línea Solumat"         / "BASE DE DATOS SOLUMAT.xlsx",
}

JT_SHEET_TO_SUBLINEA = {
    "Aeropuertos":         "aeropuertos",
    "Cargo":               "cargo_uld",
    "Cargo ULD":           "cargo_uld",
    "Carton":              "carton_corrugado",
    "Cartón":              "carton_corrugado",
    "Carton y Papel":      "carton_corrugado",
    "Cartón y Papel":      "carton_corrugado",
    "Final de Linea":      "final_linea",
    "Final de Línea":      "final_linea",
    "Motos":               "ensambladoras_motos",
    "Ensambladoras Motos": "ensambladoras_motos",
    "Solumat":             "solumat",
    "Intralogistica":      "final_linea",
    "BHS":                 "aeropuertos",
    "FINAL_LINEA":         "final_linea",
    "MOTOS":               "ensambladoras_motos",
    "AEROPUERTOS":         "aeropuertos",
    "CARTON_PAPEL":        "carton_corrugado",
    "CARGO_ULD":           "cargo_uld",
    "SOLUMAT":             "solumat",
}

if not SUPA_URL or not SUPA_KEY:
    print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY required in .env.local")
    sys.exit(1)

# ---------------------------------------------------------------------------
# pg/query HTTP client
# ---------------------------------------------------------------------------
def pg_query(sql: str) -> list:
    payload = json.dumps({"query": sql}).encode("utf-8")
    req = urllib.request.Request(
        f"{SUPA_URL}/pg/query",
        data=payload,
        headers={
            "Authorization": f"Bearer {SUPA_KEY}",
            "apikey":         SUPA_KEY,
            "Content-Type":   "application/json",
            "Content-Length": str(len(payload)),
            "User-Agent":     "MatecRadarSeed/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"pg/query HTTP {e.code}: {body[:300]}")

def pg_insert_batch(table: str, records: list, conflict_col: str | None, dry_run: bool) -> dict:
    """Bulk INSERT ... ON CONFLICT DO NOTHING (or plain INSERT) using VALUES list."""
    if not records:
        return {"inserted": 0}
    if dry_run:
        print(f"    [DRY-RUN] {len(records)} rows into {SCHEMA}.{table}")
        return {"inserted": len(records)}

    BATCH = 200
    total = 0
    for i in range(0, len(records), BATCH):
        batch = records[i:i + BATCH]
        cols = list(batch[0].keys())
        col_list = ", ".join(cols)
        values_parts = []
        for row in batch:
            vals = []
            for c in cols:
                v = row[c]
                if v is None:
                    vals.append("NULL")
                elif isinstance(v, bool):
                    vals.append("TRUE" if v else "FALSE")
                elif isinstance(v, (int, float)):
                    vals.append(str(v))
                else:
                    vals.append("'" + str(v).replace("'", "''") + "'")
            values_parts.append("(" + ", ".join(vals) + ")")

        values_sql = ",\n".join(values_parts)
        if conflict_col:
            conflict_clause = f"ON CONFLICT ({conflict_col}) DO NOTHING"
        else:
            conflict_clause = ""
        sql = f"INSERT INTO {SCHEMA}.{table} ({col_list}) VALUES\n{values_sql}\n{conflict_clause}"
        try:
            pg_query(sql)
            total += len(batch)
        except RuntimeError as e:
            print(f"    ERROR batch {i//BATCH}: {e}")
            raise

    return {"inserted": total}

# ---------------------------------------------------------------------------
# Fase 1: Sectores
# ---------------------------------------------------------------------------
def seed_sectores(dry_run: bool):
    excel_path = EXCEL_V2_FILES["aeropuertos"]
    if not excel_path.exists():
        print(f"  SKIP sectores — archivo no encontrado: {excel_path}")
        return

    print(f"\n[sectores] Leyendo {excel_path.name}...")
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "datos":
            sheet_name = name
            break
    if not sheet_name:
        for name in wb.sheetnames:
            if "dato" in name.lower() and "base" not in name.lower():
                sheet_name = name
                break

    if not sheet_name:
        print(f"  SKIP sectores — hoja 'Datos' no encontrada. Sheets: {wb.sheetnames}")
        return

    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(values_only=True))
    print(f"  Hoja '{sheet_name}': {len(rows_data)-1} filas")

    seen = set()
    records = []
    for row in rows_data[1:]:
        nombre = str(row[0]).strip() if row and row[0] else None
        if not nombre or nombre in ("None", "#N/A", "#REF!", ""):
            continue
        if nombre in seen:
            continue
        seen.add(nombre)
        records.append({"nombre": nombre, "nivel": 1, "activo": True})

    print(f"  Sectores unicos: {len(records)}")
    result = pg_insert_batch("sectores", records, None, dry_run)
    print(f"  Insertados: {result['inserted']}")

# ---------------------------------------------------------------------------
# Fase 2: Job Titles
# ---------------------------------------------------------------------------
def seed_job_titles(dry_run: bool):
    if not JOBS_FILE.exists():
        print(f"  SKIP job_titles — archivo no encontrado: {JOBS_FILE}")
        return

    print(f"\n[job_titles] Leyendo {JOBS_FILE.name}...")

    # Get sub_linea ids
    if dry_run:
        sublinea_ids = {
            "aeropuertos": 1, "cargo_uld": 2, "carton_corrugado": 3,
            "final_linea": 4, "ensambladoras_motos": 5, "solumat": 6,
        }
    else:
        rows = pg_query(f"SELECT id, codigo FROM {SCHEMA}.sub_lineas_negocio")
        sublinea_ids = {r["codigo"]: r["id"] for r in rows}
        print(f"  Sub-lineas en DB: {list(sublinea_ids.keys())}")

    wb = openpyxl.load_workbook(JOBS_FILE, read_only=True, data_only=True)
    total_inserted = 0

    def get_sublinea(sheet_name: str):
        code = JT_SHEET_TO_SUBLINEA.get(sheet_name)
        if code:
            return code
        upper = sheet_name.upper().replace("_", "").replace(" ", "")
        patterns = {
            "FINALLINEA": "final_linea", "AEROPUERTOS": "aeropuertos",
            "CARGO": "cargo_uld", "CARTONPAPEL": "carton_corrugado",
            "CARTON": "carton_corrugado", "MOTOS": "ensambladoras_motos",
            "SOLUMAT": "solumat", "CARGOULD": "cargo_uld",
        }
        for p, v in patterns.items():
            if p in upper:
                return v
        return None

    for sheet_name in wb.sheetnames:
        sublinea_codigo = get_sublinea(sheet_name)
        if not sublinea_codigo:
            print(f"  SKIP hoja '{sheet_name}' — sin mapeo")
            continue

        sub_id = sublinea_ids.get(sublinea_codigo)
        if not sub_id:
            print(f"  SKIP hoja '{sheet_name}' — sub_linea '{sublinea_codigo}' no en DB")
            continue

        ws = wb[sheet_name]
        rows_data = list(ws.iter_rows(values_only=True))
        if len(rows_data) < 2:
            continue

        header = [str(c).strip().lower() if c else "" for c in rows_data[0]]
        col = {h: i for i, h in enumerate(header)}

        titulo_col  = next((c for c in col if any(k in c for k in ["titulo", "title", "cargo"])), None)
        nivel_col   = next((c for c in col if any(k in c for k in ["nivel", "level"])), None)
        idioma_col  = next((c for c in col if any(k in c for k in ["idioma", "lang"])), None)
        prio_col    = next((c for c in col if any(k in c for k in ["prioridad", "priority"])), None)

        if not titulo_col:
            titulo_col = header[0] if header else None
        if not titulo_col:
            print(f"  SKIP hoja '{sheet_name}' — no se detecto columna de titulo")
            continue

        records = []
        seen = set()
        for row in rows_data[1:]:
            titulo = str(row[col[titulo_col]]).strip() if row[col[titulo_col]] else None
            try:
                nivel = int(row[col[nivel_col]]) if nivel_col and row[col[nivel_col]] else 3
            except (ValueError, TypeError):
                nivel = 3
            idioma = str(row[col[idioma_col]]).strip()[:2] if idioma_col and row[col[idioma_col]] else "es"
            try:
                prio = int(row[col[prio_col]]) if prio_col and row[col[prio_col]] else 2
            except (ValueError, TypeError):
                prio = 2

            if not titulo or titulo in ("None", "#N/A", ""):
                continue
            key = (sub_id, titulo.lower(), idioma)
            if key in seen:
                continue
            seen.add(key)
            records.append({
                "sub_linea_id": sub_id,
                "titulo":       titulo,
                "nivel":        min(max(nivel, 1), 5),
                "idioma":       idioma[:2],
                "prioridad":    min(max(prio, 1), 3),
                "activo":       True,
            })

        result = pg_insert_batch("job_titles_por_linea", records,
                                  "sub_linea_id, titulo, idioma", dry_run)

        print(f"  {sheet_name} -> {sublinea_codigo}: {result['inserted']} insertados")
        total_inserted += result["inserted"]

    print(f"\n  Total job_titles: {total_inserted}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--only", choices=["sectores", "job_titles"])
    args = ap.parse_args()

    print(f"Supabase: {SUPA_URL}  schema={SCHEMA}  {'[DRY-RUN]' if args.dry_run else ''}\n")

    if not args.only or args.only == "sectores":
        seed_sectores(args.dry_run)
    if not args.only or args.only == "job_titles":
        seed_job_titles(args.dry_run)

    print("\nDone.")

if __name__ == "__main__":
    main()
