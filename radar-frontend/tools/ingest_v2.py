#!/usr/bin/env python3
"""
ingest_v2.py -- Ingest empresas desde Excel/CSV a Supabase via HTTP REST API

Usa urllib.request (sin SSL, sin psycopg2). Estrategia:
  - GET empresa por company_name_norm + owner_id IS NULL
  - Si no existe -> POST (insert)
  - Si existe -> PATCH (update campos clave)
  - POST pivot empresa_sub_lineas después de cada empresa

Fuentes:
  1. V2 (93 cols): Aeropuertos, Cartón, Final de Línea, Solumat
  2. V1 (30 cols): Cargo LATAM, Ensambladoras Motos
  3. CSV: empresas_colombia_2026.csv
"""

import csv
import io
import json
import sys
import time

# Force UTF-8 output so accented characters print correctly on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl
from unidecode import unidecode

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SUPABASE_URL = "https://supabase.valparaiso.cafe"
SERVICE_KEY  = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJyb2xlIjoic2VydmljZV9yb2xlIiwiaXNzIjoic3VwYWJhc2UiLCJpYXQiOjE3MDAwMDAwMDAsImV4cCI6MjIwMDAwMDAwMH0"
    ".vBDC2y9ofT_-tvJxNdtRtFilgmkFN-ktiSl2AHiZZ3o"
)
SCHEMA       = "matec_radar"
BASE_HEADERS = {
    "apikey":          SERVICE_KEY,
    "Authorization":   f"Bearer {SERVICE_KEY}",
    "Accept-Profile":  SCHEMA,
    "Content-Profile": SCHEMA,
    "Content-Type":    "application/json",
    "User-Agent":      "MatecRadar-Frontend/2.0",
}

DOCS_DIR = Path(__file__).parent.parent.parent / "docs" / "PROSPECCIÓN"

EXCEL_FILES = [
    # (file_path, sub_linea_id, version, header_row)
    # header_row: 0-based index of the row containing column names
    (DOCS_DIR / "Linea Aeropuertos"     / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",        1, "v2", 0),
    (DOCS_DIR / "Linea Carton y Papel"  / "BASE DE DATOS CARTON Y PAPEL.xlsx",            3, "v2", 3),
    (DOCS_DIR / "Final de Línea"        / "BASE DE DATOS FINAL DE LINEA.xlsx",            4, "v2", 4),
    (DOCS_DIR / "Línea Solumat"         / "BASE DE DATOS SOLUMAT.xlsx",                   6, "v2", 3),
    (DOCS_DIR / "Línea Cargo"           / "BASE DE DATOS CARGO LATAM.xlsx",               2, "v1", 1),
    (DOCS_DIR / "Ensambladora de Motos" / "BASE DE DATOS ENSAMBLADORAS MOTOS LATAM.xlsx", 5, "v1", 0),
    (DOCS_DIR / "Línea Logistica"       / "BASE DE DATOS LOGÍSTICA 2026.xlsx",            9, "v2", 3),
]
COLOMBIA_CSV  = DOCS_DIR / "Líneas Colombianas" / "empresas_colombia_2026.csv"
CSV_SUB_LINEA = 4  # final_linea

# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------
PAIS_MAP = {
    "mexico": "MX", "méxico": "MX", "m?xico": "MX",
    "colombia": "CO",
    "brasil": "BR", "brazil": "BR",
    "argentina": "AR",
    "chile": "CL",
    "peru": "PE", "perú": "PE",
    "ecuador": "EC",
    "uruguay": "UY",
    "paraguay": "PY",
    "bolivia": "BO",
    "venezuela": "VE",
    "costa rica": "CR",
    "panama": "PA", "panamá": "PA",
    "guatemala": "GT",
    "el salvador": "SV",
    "honduras": "HN",
    "nicaragua": "NI",
    "republica dominicana": "DO", "república dominicana": "DO",
    "jamaica": "JM",
    "bahamas": "BS",
    "estados unidos": "US", "usa": "US", "united states": "US",
    "canada": "CA", "canadá": "CA",
    "espana": "ES", "españa": "ES",
}

TIER_MAP = {
    "tier a": "A", "tier_a": "A", "a": "A",
    "tier b": "B", "tier_b": "B", "b": "B",
    "tier b-alta": "B", "tier b alta": "B",
    "tier c": "C", "tier_c": "C", "c": "C",
    "tier d": "D", "tier_d": "D", "d": "D",
    "#n/a": "sin_calificar", "#ref!": "sin_calificar", "": "sin_calificar",
    "none": "sin_calificar", "sin calificar": "sin_calificar",
    "sin_calificar": "sin_calificar",
}


def norm_name(name: str) -> str:
    """unaccent + lower -- mirrors the Postgres GENERATED ALWAYS column."""
    return unidecode(str(name)).lower().strip()


def norm_pais(raw) -> tuple:
    """Returns (pais_iso, pais_nombre)."""
    if not raw or str(raw).strip() in ("None", "", "#N/A", "#N/A"):
        return None, None
    text = str(raw).strip()
    iso  = PAIS_MAP.get(text.lower()) or PAIS_MAP.get(unidecode(text).lower())
    return iso or "Otro", text


def norm_tier(raw) -> str:
    if not raw:
        return "sin_calificar"
    text = str(raw).strip()
    return TIER_MAP.get(text.lower(), "sin_calificar")


def clean_val(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "None", "nan", "", "\xa0"):
        return None
    return s


def clean_score(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return round(f, 2) if 0 <= f <= 10 else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------
GET_HEADERS = {
    "apikey":         SERVICE_KEY,
    "Authorization":  f"Bearer {SERVICE_KEY}",
    "Accept-Profile": SCHEMA,
    "User-Agent":     "MatecRadar-Frontend/2.0",
}


def http_get(path: str, params: dict | None = None) -> list | dict | None:
    url = SUPABASE_URL + "/rest/v1/" + path
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers=GET_HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        body = e.read()
        raise RuntimeError(f"GET {path} HTTP {e.code}: {body[:200]}")


def http_post(path: str, payload: dict, extra_headers: dict | None = None) -> dict | None:
    url  = SUPABASE_URL + "/rest/v1/" + path
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = dict(BASE_HEADERS)
    if extra_headers:
        headers.update(extra_headers)
    req  = urllib.request.Request(url, data=data, method="POST", headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            body = r.read()
            return json.loads(body)[0] if body else None
    except urllib.error.HTTPError as e:
        body = e.read()
        raise RuntimeError(f"POST {path} -> HTTP {e.code}: {body[:300]}")


def http_patch(path: str, params: dict, payload: dict) -> None:
    url = SUPABASE_URL + "/rest/v1/" + path + "?" + urllib.parse.urlencode(params)
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = dict(BASE_HEADERS)
    headers["Prefer"] = "return=minimal"
    req  = urllib.request.Request(url, data=data, method="PATCH", headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            r.read()
    except urllib.error.HTTPError as e:
        body = e.read()
        raise RuntimeError(f"PATCH {path} -> HTTP {e.code}: {body[:300]}")


def http_post_no_return(path: str, payload: dict) -> None:
    """POST with Prefer: return=minimal -- no response body."""
    url  = SUPABASE_URL + "/rest/v1/" + path
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = dict(BASE_HEADERS)
    headers["Prefer"] = "return=minimal"
    req  = urllib.request.Request(url, data=data, method="POST", headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            r.read()
    except urllib.error.HTTPError as e:
        body = e.read()
        raise RuntimeError(f"POST-silent {path} -> HTTP {e.code}: {body[:300]}")


# ---------------------------------------------------------------------------
# Core upsert -- GET then POST or PATCH
# ---------------------------------------------------------------------------
def upsert_empresa(empresa: dict) -> tuple[int | None, str]:
    """
    Returns (empresa_id, action) where action is 'insert', 'update', or 'skip'.
    """
    name = (empresa.get("company_name") or "").strip()
    if not name:
        return None, "skip"

    name_norm = norm_name(name)

    # Look up existing
    rows = http_get(
        "empresas",
        {"company_name_norm": f"eq.{name_norm}", "owner_id": "is.null", "select": "id"}
    )

    if rows:
        emp_id = rows[0]["id"]
        # PATCH to update key fields
        update_payload = {}
        for k in ("pais", "pais_nombre", "industria_cliente", "grupo_empresarial",
                   "company_domain", "ciudad", "estado_region", "tier_actual",
                   "score_total_ultimo", "meta", "source_file", "source_sheet"):
            if empresa.get(k) is not None:
                update_payload[k] = empresa[k]
        if update_payload:
            http_patch(
                "empresas",
                {"id": f"eq.{emp_id}", "owner_id": "is.null"},
                update_payload
            )
        return emp_id, "update"
    else:
        # POST to insert
        payload = {
            "company_name":      name,
            "pais":              empresa.get("pais"),
            "pais_nombre":       empresa.get("pais_nombre"),
            "estado_region":     empresa.get("estado_region"),
            "ciudad":            empresa.get("ciudad"),
            "industria_cliente": empresa.get("industria_cliente"),
            "grupo_empresarial": empresa.get("grupo_empresarial"),
            "company_domain":    empresa.get("company_domain"),
            "tier_actual":       empresa.get("tier_actual", "sin_calificar"),
            "score_total_ultimo": empresa.get("score_total_ultimo"),
            "prioridad":         "media",
            "pipeline":          "no_iniciado",
            "meta":              empresa.get("meta", {}),
            "source_file":       empresa.get("source_file"),
            "source_sheet":      empresa.get("source_sheet"),
        }
        # Remove None values so Postgres defaults apply
        payload = {k: v for k, v in payload.items() if v is not None}

        headers = {"Prefer": "return=representation"}
        row = http_post("empresas", payload, extra_headers=headers)
        if row and "id" in row:
            return row["id"], "insert"
        return None, "skip"


def upsert_pivot(empresa_id: int, sub_linea_id: int, es_principal: bool = True) -> None:
    """Insert pivot row empresa_sub_lineas -- DO NOTHING on conflict."""
    try:
        http_post_no_return(
            "empresa_sub_lineas",
            {
                "empresa_id":   empresa_id,
                "sub_linea_id": sub_linea_id,
                "es_principal": es_principal,
            }
        )
    except RuntimeError as e:
        # 409 conflict means already exists -- that's fine
        if "409" not in str(e) and "23505" not in str(e):
            raise


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
def load_sheet(path: Path, sheet_name: str = "Base de Datos", header_row: int = 0):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target = None
    for name in wb.sheetnames:
        if sheet_name.lower() in name.lower():
            target = name
            break
    if not target:
        target = wb.sheetnames[0]
    ws   = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(c).strip() if c is not None else f"col_{i}"
              for i, c in enumerate(rows[header_row])]
    return header, rows[header_row + 1:]


def find_col(header: list, *candidates) -> int | None:
    """Return first column index matching any candidate substring (case-insensitive)."""
    for cand in candidates:
        cand_norm = unidecode(cand).lower()
        for i, h in enumerate(header):
            if cand_norm in unidecode(h).lower():
                return i
    return None


def gcol(row, header, *candidates):
    """Get cleaned value from row by column name candidates."""
    idx = find_col(header, *candidates)
    if idx is None or idx >= len(row):
        return None
    return clean_val(row[idx])


# ---------------------------------------------------------------------------
# V2 parser (93 cols) -- Aeropuertos, Cartón, Final de Línea, Solumat
# ---------------------------------------------------------------------------
def parse_v2(row, header, source_file: str) -> dict:
    pais_raw  = gcol(row, header, "PAIS", "País", "PA?S", "Country")
    pais_iso, pais_nombre = norm_pais(pais_raw)

    tier_raw  = gcol(row, header, "TIER")
    score_raw = gcol(row, header, "Score_Total", "Score Total")

    # company name: search by column header (some files have col[0]=None, col[1]=ID, col[2]=name)
    company_name = gcol(row, header, "COMPANY NAME", "Empresa", "Company")

    # grupo empresarial: prefer col index 2 if header contains "Grupo" there,
    # else search by name
    grupo = gcol(row, header, "GRUPO EMPRESARIAL", "Grupo_Empresarial", "Grupo Empresarial", "Holding")
    dominio = gcol(row, header, "DOMINIO", "Domain", "Web", "URL")
    industria = gcol(row, header, "INDSUTRIA", "INDUSTRIA", "Industria", "Sector")

    # Meta: scoring fields + radar fields
    meta_keys = [
        ("impacto",     "IMPACTO EN EL PRESUPUESTO", "Impacto"),
        ("multiplanta", "MULTIPLANTA", "Multiplanta"),
        ("recurrencia", "RECURRENCIA", "Recurrencia"),
        ("referente",   "REFERENTE DEL MERCADO", "Referente"),
        ("anio_obj",    "A?O OBJETIVO", "Año Objetivo"),
        ("prioridad_c", "PRIORIDAD", "Prioridad"),
        ("ticket",      "TICKET ESTIMADO", "Ticket"),
        ("ventana",     "VENTANA DE COMPRA", "Ventana"),
        ("tipo_senal",  "TIPO DE SE?AL", "Tipo de Señal"),
        ("pipeline",    "PIPELINE"),
        ("semaforo",    "SEMAFORO", "SEM?FORO"),
    ]
    meta = {}
    for key, *cands in meta_keys:
        val = gcol(row, header, *cands)
        if val:
            meta[key] = val

    return {
        "company_name":      company_name or "",
        "pais":              pais_iso,
        "pais_nombre":       pais_nombre,
        "estado_region":     gcol(row, header, "ESTADO", "DEPARTAMENTO", "Departamento"),
        "ciudad":            gcol(row, header, "CIUDAD", "Ciudad", "City"),
        "industria_cliente": industria,
        "grupo_empresarial": grupo,
        "company_domain":    dominio,
        "tier_actual":       norm_tier(tier_raw),
        "score_total_ultimo": clean_score(score_raw),
        "source_file":       source_file,
        "source_sheet":      "Base de Datos",
        "meta":              meta,
    }


# ---------------------------------------------------------------------------
# V1 parser (30 cols) -- Cargo, Motos
# ---------------------------------------------------------------------------
def parse_v1(row, header, source_file: str) -> dict:
    pais_raw  = gcol(row, header, "Pais", "País", "Country")
    pais_iso, pais_nombre = norm_pais(pais_raw)

    tier_raw  = gcol(row, header, "TIER_A_B_C", "TIER", "Tier")
    score_raw = gcol(row, header, "Score_Total", "Score")

    meta = {}
    for key in ("Tipo_Operacion", "Linea_Negocio_Matec", "Cuenta_Estrategica_SI_NO",
                "Prioridad_Comercial", "Radar_Activo_SI_NO", "Tipo_Senal_Inversion",
                "Ventana_Compra", "Dolor_Principal_Cliente"):
        val = gcol(row, header, key)
        if val:
            meta[key.lower()] = val

    return {
        "company_name":      gcol(row, header, "COMPANY NAME", "Empresa", "Company") or "",
        "pais":              pais_iso,
        "pais_nombre":       pais_nombre,
        "ciudad":            gcol(row, header, "Ciudad", "City"),
        "grupo_empresarial": gcol(row, header, "Grupo_Empresarial", "Grupo"),
        "industria_cliente": gcol(row, header, "Industria_Cliente", "Industria", "Sector"),
        "company_domain":    gcol(row, header, "URL_Web", "Dominio", "Web"),
        "tier_actual":       norm_tier(tier_raw),
        "score_total_ultimo": clean_score(score_raw),
        "source_file":       source_file,
        "source_sheet":      "Base de Datos",
        "meta":              meta,
    }


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------
class Stats:
    def __init__(self, label: str):
        self.label    = label
        self.inserted = 0
        self.updated  = 0
        self.skipped  = 0
        self.errors   = 0

    def report(self):
        print(
            f"  {self.label:50s}  "
            f"ins={self.inserted:5d}  upd={self.updated:5d}  "
            f"skip={self.skipped:5d}  err={self.errors:5d}"
        )


all_stats: list[Stats] = []


# ---------------------------------------------------------------------------
# Runner for Excel files
# ---------------------------------------------------------------------------
def run_excel(path: Path, sub_linea_id: int, version: str, header_row: int = 0):
    label = f"{path.name} -> sub_linea_id={sub_linea_id}"
    st    = Stats(label)
    all_stats.append(st)

    if not path.exists():
        print(f"  SKIP (not found): {path}")
        st.skipped += 1
        return

    print(f"\n[{version.upper()}] {path.name} -> sub_linea_id={sub_linea_id} (header_row={header_row})")
    header, rows = load_sheet(path, header_row=header_row)
    if not header:
        print("  SKIP: could not read sheet")
        st.skipped += 1
        return

    parse_fn = parse_v2 if version == "v2" else parse_v1

    for i, row in enumerate(rows):
        if not any(row):
            continue
        try:
            empresa = parse_fn(row, header, path.name)
            if not empresa.get("company_name"):
                st.skipped += 1
                continue

            emp_id, action = upsert_empresa(empresa)

            if emp_id is None:
                st.skipped += 1
                continue

            upsert_pivot(emp_id, sub_linea_id)

            if action == "insert":
                st.inserted += 1
            else:
                st.updated += 1

            if (st.inserted + st.updated) % 25 == 0:
                print(f"    row {i+1}: ins={st.inserted} upd={st.updated} err={st.errors}")

        except Exception as e:
            st.errors += 1
            name_val = row[1] if len(row) > 1 else "?"
            print(f"  ERROR row {i+1} ({name_val}): {e}")
            # Small back-off in case of transient error
            time.sleep(0.1)

    print(f"  Done: ins={st.inserted} upd={st.updated} skip={st.skipped} err={st.errors}")


# ---------------------------------------------------------------------------
# Runner for Colombia CSV
# ---------------------------------------------------------------------------
def run_csv():
    label = f"{COLOMBIA_CSV.name} -> sub_linea_id={CSV_SUB_LINEA}"
    st    = Stats(label)
    all_stats.append(st)

    if not COLOMBIA_CSV.exists():
        print(f"  SKIP (not found): {COLOMBIA_CSV}")
        st.skipped += 1
        return

    print(f"\n[CSV] {COLOMBIA_CSV.name} -> sub_linea_id={CSV_SUB_LINEA}")

    with open(COLOMBIA_CSV, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row_raw in enumerate(reader):
            try:
                name = (row_raw.get("COMPANY NAME") or "").strip()
                if not name:
                    st.skipped += 1
                    continue

                pais_raw = row_raw.get("pais") or "CO"
                pais_iso, pais_nombre = norm_pais(pais_raw)

                empresa = {
                    "company_name":      name,
                    "pais":              pais_iso or "CO",
                    "pais_nombre":       pais_nombre or "Colombia",
                    "industria_cliente": clean_val(row_raw.get("sector")),
                    "tier_actual":       "sin_calificar",
                    "source_file":       COLOMBIA_CSV.name,
                    "source_sheet":      "csv",
                    "meta": {
                        "source_csv":  (row_raw.get("source") or "").upper(),
                        "fuente_col":  True,
                    },
                }

                emp_id, action = upsert_empresa(empresa)

                if emp_id is None:
                    st.skipped += 1
                    continue

                upsert_pivot(emp_id, CSV_SUB_LINEA)

                if action == "insert":
                    st.inserted += 1
                else:
                    st.updated += 1

                if (st.inserted + st.updated) % 25 == 0:
                    print(f"    row {i+1}: ins={st.inserted} upd={st.updated} err={st.errors}")

            except Exception as e:
                st.errors += 1
                print(f"  ERROR CSV row {i+1}: {e}")
                time.sleep(0.1)

    print(f"  Done: ins={st.inserted} upd={st.updated} skip={st.skipped} err={st.errors}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 65)
    print("  ingest_v2.py -- Excel/CSV -> Supabase (HTTP API)")
    print(f"  Target: {SUPABASE_URL}  schema={SCHEMA}")
    print("=" * 65)

    # Verify connection
    try:
        rows = http_get("sub_lineas_negocio", {"select": "id,codigo"})
        print(f"\nConnection OK -- sub_lineas: {rows}\n")
    except Exception as e:
        print(f"ERROR connecting to Supabase: {e}")
        sys.exit(1)

    # Run all Excel files
    for file_path, sub_linea_id, version, header_row in EXCEL_FILES:
        run_excel(file_path, sub_linea_id, version, header_row)

    # Run CSV
    run_csv()

    # Final report
    print("\n" + "=" * 65)
    print("  REPORTE FINAL")
    print("=" * 65)
    total_ins = total_upd = total_skip = total_err = 0
    for st in all_stats:
        st.report()
        total_ins  += st.inserted
        total_upd  += st.updated
        total_skip += st.skipped
        total_err  += st.errors

    print("-" * 65)
    print(
        f"  {'TOTAL':50s}  "
        f"ins={total_ins:5d}  upd={total_upd:5d}  "
        f"skip={total_skip:5d}  err={total_err:5d}"
    )
    if total_err:
        print(f"\n  WARNING: {total_err} errores -- revisar output arriba")
    else:
        print("\n  OK -- sin errores")


if __name__ == "__main__":
    main()
