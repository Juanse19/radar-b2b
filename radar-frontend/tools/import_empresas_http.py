#!/usr/bin/env python3
"""
import_empresas_http.py
Import empresas from Excel/CSV to Supabase via HTTP /pg/query.
No requires SUPABASE_DB_URL — only SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY.

Usage:
  pip install openpyxl python-dotenv unidecode
  python radar-frontend/tools/import_empresas_http.py [--dry-run] [--skip-errors]
    [--only aeropuertos|carton_corrugado|final_linea|solumat|cargo_uld|ensambladoras_motos|colombia]
"""

import os, sys, csv, json, argparse
from pathlib import Path

from dotenv import load_dotenv
env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)
env_local = Path(__file__).parent.parent / ".env.local"
if env_local.exists():
    load_dotenv(env_local, override=True)

try:
    import openpyxl
    from unidecode import unidecode
    import urllib.request, urllib.error
except ImportError as e:
    print(f"ERROR: missing dependency — {e}")
    print("Install: pip install openpyxl python-dotenv unidecode")
    sys.exit(1)

SUPA_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPA_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SCHEMA   = "matec_radar"
REPO_ROOT = Path(__file__).parent.parent.parent
DOCS_DIR  = REPO_ROOT / "docs" / "PROSPECCIÓN"

if not SUPA_URL or not SUPA_KEY:
    print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY required in .env.local")
    sys.exit(1)

EXCEL_FILES = [
    {"file": DOCS_DIR / "Linea Aeropuertos" / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",
     "sub_linea": "aeropuertos", "version": "v2"},
    {"file": DOCS_DIR / "Linea Carton y Papel" / "BASE DE DATOS CARTON Y PAPEL.xlsx",
     "sub_linea": "carton_corrugado", "version": "v2"},
    {"file": DOCS_DIR / "Final de Línea" / "BASE DE DATOS FINAL DE LINEA.xlsx",
     "sub_linea": "final_linea", "version": "v2"},
    {"file": DOCS_DIR / "Línea Solumat" / "BASE DE DATOS SOLUMAT.xlsx",
     "sub_linea": "solumat", "version": "v2"},
    {"file": DOCS_DIR / "Línea Cargo" / "BASE DE DATOS CARGO LATAM.xlsx",
     "sub_linea": "cargo_uld", "version": "v1"},
    {"file": DOCS_DIR / "Ensambladora de Motos" / "BASE DE DATOS ENSAMBLADORAS MOTOS LATAM.xlsx",
     "sub_linea": "ensambladoras_motos", "version": "v1"},
]
COLOMBIA_CSV = DOCS_DIR / "Líneas Colombianas" / "empresas_colombia_2026.csv"

# CSV sub_linea mapping for Colombian companies
COLOMBIA_LINEA_MAP = {
    "ALIMENTICIO": "final_linea",
    "ALIMENTOS":   "final_linea",
    "CARTON":      "carton_corrugado",
    "PAPEL":       "carton_corrugado",
    "MOTOS":       "ensambladoras_motos",
    "BHS":         "aeropuertos",
    "AEROPUERTOS": "aeropuertos",
}

# ── Normalization ─────────────────────────────────────────────────────────────

PAIS_MAP = {
    "mexico":"MX","méxico":"MX","colombia":"CO","brasil":"BR","brazil":"BR",
    "argentina":"AR","chile":"CL","peru":"PE","perú":"PE","ecuador":"EC",
    "uruguay":"UY","paraguay":"PY","bolivia":"BO","venezuela":"VE",
    "costa rica":"CR","panama":"PA","panamá":"PA","guatemala":"GT",
    "el salvador":"SV","honduras":"HN","nicaragua":"NI",
    "republica dominicana":"DO","república dominicana":"DO",
    "jamaica":"JM","bahamas":"BS","estados unidos":"US","usa":"US",
    "canada":"CA","canadá":"CA","espana":"ES","españa":"ES",
}
TIER_MAP = {
    "tier a":"A","tier_a":"A","a":"A",
    "tier b":"B","tier_b":"B","b":"B","tier b-alta":"B","tier b alta":"B",
    "tier c":"C","tier_c":"C","c":"C",
    "tier d":"D","tier_d":"D","d":"D",
    "#n/a":"sin_calificar","#ref!":"sin_calificar","":"sin_calificar",
    "none":"sin_calificar","sin calificar":"sin_calificar","sin_calificar":"sin_calificar",
}

def norm_pais(raw):
    if not raw or str(raw).strip() in ("None","","#N/A"): return None, None
    text = str(raw).strip()
    iso = PAIS_MAP.get(text.lower()) or PAIS_MAP.get(unidecode(text).lower())
    return (iso or "Otro"), text

def norm_tier(raw):
    if not raw: return "sin_calificar"
    return TIER_MAP.get(str(raw).strip().lower(), "sin_calificar")

def clean_val(val):
    if val is None: return None
    s = str(val).strip()
    if s in ("#N/A","#REF!","#VALUE!","#DIV/0!","#NAME?","None","nan",""): return None
    return s

def clean_score(val):
    if val is None: return None
    try:
        f = float(val)
        return round(f, 2) if 0 <= f <= 10 else None
    except (ValueError, TypeError):
        return None

# ── HTTP client ───────────────────────────────────────────────────────────────

def pg_query(sql: str) -> list:
    payload = json.dumps({"query": sql}).encode("utf-8")
    req = urllib.request.Request(
        f"{SUPA_URL}/pg/query",
        data=payload,
        headers={
            "Authorization": f"Bearer {SUPA_KEY}",
            "apikey": SUPA_KEY,
            "Content-Type": "application/json",
            "Content-Length": str(len(payload)),
            "User-Agent": "MatecRadarImport/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"pg/query HTTP {e.code}: {body[:400]}")

# ── SQL value quoting ─────────────────────────────────────────────────────────

def q(v):
    """Quote a Python value for SQL embedding."""
    if v is None: return "NULL"
    if isinstance(v, bool): return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)): return str(v)
    if isinstance(v, dict): return "'" + json.dumps(v, ensure_ascii=False).replace("'","''") + "'::jsonb"
    return "'" + str(v).replace("'","''") + "'"

# ── Excel parsing ─────────────────────────────────────────────────────────────

def load_sheet(path: Path, sheet_name: str = "Base de Datos"):
    if not path.exists():
        print(f"  SKIP — file not found: {path}")
        return [], []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target = None
    for name in wb.sheetnames:
        if sheet_name.lower() in name.lower():
            target = name; break
    if not target:
        target = wb.sheetnames[0]
    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return [], []
    header = [str(c).strip() if c else f"col_{i}" for i,c in enumerate(rows[0])]
    return header, rows[1:]

def parse_v2(row, header, sub_linea: str, source: str) -> dict | None:
    col = {h: i for i,h in enumerate(header)}
    def g(*keys):
        for k in keys:
            for col_name, idx in col.items():
                if k.lower() in col_name.lower():
                    v = clean_val(row[idx] if idx < len(row) else None)
                    if v: return v
        return None

    name = g("Empresa","Company","Nombre_Empresa","Nombre Empresa") or ""
    if not name: return None
    pais_iso, pais_nombre = norm_pais(g("Pais","País","Country"))
    return {
        "company_name":      name,
        "pais":              pais_iso,
        "pais_nombre":       pais_nombre,
        "estado_region":     g("Estado","Region","Región","Departamento"),
        "ciudad":            g("Ciudad","City"),
        "industria_cliente": g("Industria","Sector","Giro"),
        "grupo_empresarial": g("Grupo","Grupo Empresarial","Holding"),
        "company_domain":    g("Dominio","Domain","Web","URL"),
        "company_url":       g("URL","Website","Sitio Web"),
        "tier_actual":       norm_tier(g("Tier","TIER","Clasificación")),
        "score_total_ultimo": clean_score(g("Score","SCORE","Puntaje","Score_Total")),
        "source_file":       source,
        "source_sheet":      "Base de Datos",
        "meta": {"sub_linea_codigo": sub_linea, "meta_schema": "v2_amplio"},
        "_sub_linea": sub_linea,
    }

def parse_v1(row, header, sub_linea: str, source: str) -> dict | None:
    col = {h.strip(): i for i,h in enumerate(header)}
    def g(*keys):
        for k in keys:
            for col_name, idx in col.items():
                if k.lower() in col_name.lower():
                    v = clean_val(row[idx] if idx < len(row) else None)
                    if v: return v
        return None

    name = g("Empresa","Company","Nombre") or ""
    if not name: return None
    pais_iso, pais_nombre = norm_pais(g("Pais","País","Country"))
    return {
        "company_name":      name,
        "pais":              pais_iso,
        "pais_nombre":       pais_nombre,
        "ciudad":            g("Ciudad","City"),
        "industria_cliente": g("Industria","Sector","Giro"),
        "grupo_empresarial": g("Grupo","Holding"),
        "company_domain":    g("Dominio","Web","URL"),
        "tier_actual":       norm_tier(g("Tier","TIER")),
        "score_total_ultimo": clean_score(g("Score","Puntaje")),
        "source_file":       source,
        "source_sheet":      "Base de Datos",
        "meta": {"sub_linea_codigo": sub_linea, "meta_schema": "v1_compacto"},
        "_sub_linea": sub_linea,
    }

def parse_colombia_csv(path: Path) -> list[dict]:
    if not path.exists():
        print(f"  SKIP — CSV not found: {path}")
        return []
    records = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Support both "COMPANY NAME" and "empresa"/"Empresa"/"company_name"
            name = (
                row.get("COMPANY NAME") or row.get("company name") or
                row.get("empresa") or row.get("Empresa") or row.get("company_name") or ""
            ).strip()
            if not name: continue
            # Determine sub_linea from 'linea' column or fallback to solumat for Colombian plastics
            linea_raw = (row.get("linea") or row.get("linea_negocio") or row.get("sector") or "").strip().upper()
            sub_linea = "solumat"  # default for Colombian plastics/packaging companies
            for key, val in COLOMBIA_LINEA_MAP.items():
                if key in linea_raw:
                    sub_linea = val
                    break
            pais_iso, pais_nombre = norm_pais(row.get("pais") or "Colombia")
            records.append({
                "company_name":      name,
                "pais":              pais_iso or "CO",
                "pais_nombre":       pais_nombre or "Colombia",
                "ciudad":            clean_val(row.get("ciudad")),
                "industria_cliente": clean_val(row.get("sector") or row.get("industria")),
                "company_domain":    clean_val(row.get("dominio") or row.get("web")),
                "tier_actual":       "sin_calificar",
                "source_file":       path.name,
                "source_sheet":      "colombia_2026",
                "meta": {"sub_linea_codigo": sub_linea, "meta_schema": "v1_compacto"},
                "_sub_linea": sub_linea,
            })
    return records

# ── Insert logic ──────────────────────────────────────────────────────────────

# Columns we'll insert (company_name_norm is GENERATED — do NOT include)
INSERT_COLS = [
    "company_name","pais","pais_nombre","estado_region","ciudad",
    "industria_cliente","grupo_empresarial","company_domain","company_url",
    "tier_actual","score_total_ultimo","source_file","source_sheet","meta",
]

def insert_empresas_batch(records: list[dict], dry_run: bool, skip_errors: bool) -> dict[str, int]:
    """
    INSERT records in batches of 100.
    Returns {company_name_normalized: empresa_id} for all inserted/existing rows.
    """
    if not records: return {}
    BATCH = 100
    for i in range(0, len(records), BATCH):
        batch = records[i:i+BATCH]
        rows_sql = []
        for rec in batch:
            vals = []
            for col in INSERT_COLS:
                vals.append(q(rec.get(col)))
            rows_sql.append("(" + ", ".join(vals) + ")")
        col_list = ", ".join(INSERT_COLS)
        sql = (
            f"INSERT INTO {SCHEMA}.empresas ({col_list}) VALUES\n"
            + ",\n".join(rows_sql)
            + "\nON CONFLICT (company_name_norm) WHERE owner_id IS NULL DO UPDATE SET"
            + "\n  pais = EXCLUDED.pais,"
            + "\n  pais_nombre = EXCLUDED.pais_nombre,"
            + "\n  industria_cliente = EXCLUDED.industria_cliente,"
            + "\n  meta = EXCLUDED.meta,"
            + "\n  source_file = EXCLUDED.source_file,"
            + "\n  imported_at = NOW(),"
            + "\n  updated_at = NOW()"
        )
        if dry_run:
            print(f"  [DRY-RUN] Would insert/upsert {len(batch)} empresas")
            continue
        try:
            pg_query(sql)
            print(f"  Batch {i//BATCH + 1}: {len(batch)} empresas upserted")
        except RuntimeError as e:
            print(f"  ERROR batch {i//BATCH + 1}: {e}")
            if not skip_errors:
                raise

    # Now SELECT back all IDs for these company names
    if dry_run:
        return {}
    names = [r["company_name"] for r in records]
    CHUNK = 200
    id_map: dict[str, int] = {}
    for i in range(0, len(names), CHUNK):
        chunk = names[i:i+CHUNK]
        arr = ", ".join("'" + n.replace("'","''") + "'" for n in chunk)
        rows = pg_query(
            f"SELECT id, company_name FROM {SCHEMA}.empresas "
            f"WHERE company_name = ANY(ARRAY[{arr}]) AND owner_id IS NULL"
        )
        for row in rows:
            id_map[row["company_name"]] = row["id"]
    return id_map

def insert_sub_lineas_batch(links: list[tuple[int,int]], dry_run: bool, skip_errors: bool):
    """Insert (empresa_id, sub_linea_id) pairs into empresa_sub_lineas."""
    if not links or dry_run:
        if dry_run: print(f"  [DRY-RUN] Would insert {len(links)} empresa_sub_lineas links")
        return
    BATCH = 500
    for i in range(0, len(links), BATCH):
        batch = links[i:i+BATCH]
        vals = ", ".join(f"({eid}, {slid}, TRUE)" for eid, slid in batch)
        sql = (
            f"INSERT INTO {SCHEMA}.empresa_sub_lineas (empresa_id, sub_linea_id, es_principal) VALUES {vals} "
            f"ON CONFLICT (empresa_id, sub_linea_id) DO NOTHING"
        )
        try:
            pg_query(sql)
        except RuntimeError as e:
            print(f"  ERROR empresa_sub_lineas batch: {e}")
            if not skip_errors: raise

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-errors", action="store_true")
    ap.add_argument("--only", help="Import only this sub_linea (or 'colombia')")
    args = ap.parse_args()

    # Fetch sub_linea ID map from production
    print("Fetching sub_linea IDs from production...")
    if not args.dry_run:
        rows = pg_query(f"SELECT id, codigo FROM {SCHEMA}.sub_lineas_negocio")
        sublinea_ids: dict[str, int] = {r["codigo"]: r["id"] for r in rows}
        print(f"  Sub-lineas: {list(sublinea_ids.keys())}")
    else:
        sublinea_ids = {
            "aeropuertos": 1, "cargo_uld": 2, "carton_corrugado": 3,
            "final_linea": 4, "ensambladoras_motos": 5, "solumat": 6,
        }

    total_empresas = 0
    total_links    = 0

    # ── Excel files ───────────────────────────────────────────────────────────
    for entry in EXCEL_FILES:
        sub_linea = entry["sub_linea"]
        if args.only and args.only != sub_linea: continue

        sub_linea_id = sublinea_ids.get(sub_linea)
        if not sub_linea_id:
            print(f"\nSKIP {sub_linea} — not found in sub_lineas_negocio")
            continue

        print(f"\n[{sub_linea}] Reading {entry['file'].name}...")
        header, rows = load_sheet(entry["file"])
        if not rows:
            print(f"  SKIP — empty or not found")
            continue

        parse_fn = parse_v2 if entry["version"] == "v2" else parse_v1
        records: list[dict] = []
        seen = set()
        for row in rows:
            emp = parse_fn(row, header, sub_linea, entry["file"].name)
            if not emp or not emp["company_name"]: continue
            norm = unidecode(emp["company_name"]).lower().strip()
            if norm in seen: continue
            seen.add(norm)
            records.append(emp)

        print(f"  Parsed: {len(records)} unique empresas")
        if not records: continue

        id_map = insert_empresas_batch(records, args.dry_run, args.skip_errors)
        total_empresas += len(records)

        # Build empresa_sub_lineas links
        links: list[tuple[int,int]] = []
        for emp in records:
            eid = id_map.get(emp["company_name"])
            if eid:
                links.append((eid, sub_linea_id))
        print(f"  Linking {len(links)} empresas to sub_linea '{sub_linea}'")
        insert_sub_lineas_batch(links, args.dry_run, args.skip_errors)
        total_links += len(links)

    # ── Colombia CSV ─────────────────────────────────────────────────────────
    if not args.only or args.only == "colombia":
        print(f"\n[colombia] Reading {COLOMBIA_CSV.name}...")
        colombia_records = parse_colombia_csv(COLOMBIA_CSV)
        print(f"  Parsed: {len(colombia_records)} records")

        # Group by sub_linea for linking
        by_sublinea: dict[str, list[dict]] = {}
        for emp in colombia_records:
            sl = emp["_sub_linea"]
            by_sublinea.setdefault(sl, []).append(emp)

        # Remove duplicates across all Colombia records
        seen = set()
        unique_records = []
        for emp in colombia_records:
            norm = unidecode(emp["company_name"]).lower().strip()
            if norm in seen: continue
            seen.add(norm)
            unique_records.append(emp)

        print(f"  Unique: {len(unique_records)}")
        id_map = insert_empresas_batch(unique_records, args.dry_run, args.skip_errors)
        total_empresas += len(unique_records)

        for sl, emps in by_sublinea.items():
            sl_id = sublinea_ids.get(sl)
            if not sl_id:
                print(f"  SKIP sub_linea '{sl}' for Colombia — not in DB")
                continue
            links = [(id_map[e["company_name"]], sl_id)
                     for e in emps if e["company_name"] in id_map]
            insert_sub_lineas_batch(links, args.dry_run, args.skip_errors)
            total_links += len(links)

    print(f"\n{'='*60}")
    print(f"  TOTAL empresas processed : {total_empresas}")
    print(f"  TOTAL sub_linea links    : {total_links}")
    print(f"  Mode: {'DRY-RUN' if args.dry_run else 'LIVE'}")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
