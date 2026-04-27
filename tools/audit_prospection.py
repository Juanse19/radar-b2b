"""
Matec Radar B2B — Audit: Excel Prospection Files vs Supabase Production
Read-only. No DB writes.
"""

import unicodedata
import requests
import openpyxl
from pathlib import Path
from collections import defaultdict

# ─── CONFIG ──────────────────────────────────────────────────────────────────

BASE_PATH = Path("c:/Users/Juan/Documents/Agentic Workflows/clients/docs/PROSPECCIÓN")

# Per-file config after manual inspection:
# - header_row: 0-indexed row containing column names
# - col_name:   exact header string of the company name column
# - sheet:      sheet name to use (None = first sheet)
EXCEL_FILES = [
    {
        "sub_linea": "aeropuertos",
        "path": BASE_PATH / "Linea Aeropuertos" / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",
        "sheet": "Base de Datos",
        "header_row": 0,   # row 0 has ID, COMPANY NAME, ...
        "col_name": "COMPANY NAME",
    },
    {
        "sub_linea": "cargo_uld",
        "path": BASE_PATH / "Línea Cargo" / "BASE DE DATOS CARGO LATAM.xlsx",
        "sheet": "CRM_Cargo_ULD_LATAM",  # first sheet, header at row 1 (row 0 is blank/merge)
        "header_row": 1,
        "col_name": "Empresa",
    },
    {
        "sub_linea": "carton_corrugado",
        "path": BASE_PATH / "Linea Carton y Papel" / "BASE DE DATOS CARTON Y PAPEL.xlsx",
        "sheet": "Base de Datos",
        "header_row": 3,   # rows 0-2 are category/weight labels; row 3 = ID, COMPANY NAME, ...
        "col_name": "COMPANY NAME",
    },
    {
        "sub_linea": "final_linea",
        "path": BASE_PATH / "Final de Línea" / "BASE DE DATOS FINAL DE LINEA.xlsx",
        "sheet": "Base de Datos",
        "header_row": 4,   # rows 0-3 are merged/category rows; row 4 = ID, COMPANY NAME, ...
        "col_name": "COMPANY NAME",
    },
    {
        "sub_linea": "ensambladoras_motos",
        "path": BASE_PATH / "Ensambladora de Motos" / "BASE DE DATOS ENSAMBLADORAS MOTOS LATAM.xlsx",
        "sheet": "Base de Datos",
        "header_row": 0,
        "col_name": "COMPANY NAME",
    },
    {
        "sub_linea": "solumat",
        "path": BASE_PATH / "Línea Solumat" / "BASE DE DATOS SOLUMAT.xlsx",
        "sheet": "Base de Datos",
        "header_row": 3,   # rows 0-2 are category/weight labels; row 3 = ID, COMPANY NAME, ...
        "col_name": "COMPANY NAME",
    },
    {
        "sub_linea": "logistica",
        "path": BASE_PATH / "Línea Logistica" / "BASE DE DATOS LOGÍSTICA 2026.xlsx",
        "sheet": "Base de Datos",
        "header_row": 3,   # row 3 = ID, COMPANY NAME, ...
        "col_name": "COMPANY NAME",
    },
]

SUPABASE_URL = "https://supabase.valparaiso.cafe"
SERVICE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJyb2xlIjoic2VydmljZV9yb2xlIiwiaXNzIjoic3VwYWJhc2UiLCJpYXQiOjE3MDAwMDAwMDAsImV4cCI6MjIwMDAwMDAwMH0"
    ".vBDC2y9ofT_-tvJxNdtRtFilgmkFN-ktiSl2AHiZZ3o"
)
SCHEMA = "matec_radar"

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Accept-Profile": SCHEMA,
}

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def norm_name(s: str) -> str:
    """Normalize company name: NFD→ASCII, lowercase, strip."""
    if not s:
        return ""
    s = str(s).strip()
    nfd = unicodedata.normalize("NFD", s)
    ascii_str = nfd.encode("ascii", "ignore").decode("ascii")
    return ascii_str.lower().strip()


def find_col(headers: list, target: str) -> int | None:
    target_norm = target.strip().lower()
    for i, h in enumerate(headers):
        if h and str(h).strip().lower() == target_norm:
            return i
    return None


def supabase_get(path: str, params: dict = None) -> tuple:
    """GET from Supabase. Returns (data, total_count, error)."""
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    hdrs = dict(HEADERS)
    hdrs["Prefer"] = "count=exact"
    try:
        resp = requests.get(url, headers=hdrs, params=params, timeout=30)
        total = None
        cr = resp.headers.get("Content-Range", "")
        if "/" in cr:
            try:
                total = int(cr.split("/")[1])
            except ValueError:
                pass
        if resp.status_code in (200, 206):
            return resp.json(), total, None
        else:
            return None, None, f"HTTP {resp.status_code}: {resp.text[:300]}"
    except Exception as e:
        return None, None, str(e)


# ─── STEP 1: EXCEL COUNTS ────────────────────────────────────────────────────

print("=" * 70)
print("STEP 1: READING EXCEL FILES")
print("=" * 70)

excel_data = {}

for cfg in EXCEL_FILES:
    sub = cfg["sub_linea"]
    path = cfg["path"]
    hdr_row_idx = cfg["header_row"]
    sheet_name = cfg["sheet"]
    col_name = cfg["col_name"]

    print(f"\n[{sub}] {path.name}")

    if not path.exists():
        print(f"  ERROR: File not found at {path}")
        excel_data[sub] = {"error": "file not found"}
        continue

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]
            print(f"  WARNING: sheet '{sheet_name}' not found, using '{ws.title}'")

        all_rows = list(ws.iter_rows(values_only=True))
        print(f"  Sheet rows: {len(all_rows)}, header_row={hdr_row_idx}, looking for col='{col_name}'")

        if len(all_rows) <= hdr_row_idx:
            print(f"  ERROR: header_row={hdr_row_idx} out of range ({len(all_rows)} rows)")
            excel_data[sub] = {"error": "header row out of range"}
            wb.close()
            continue

        header_row = all_rows[hdr_row_idx]
        col_idx = find_col(list(header_row), col_name)

        if col_idx is None:
            avail = [str(h).strip() for h in header_row if h and str(h).strip()]
            print(f"  WARNING: '{col_name}' not found. Available: {avail[:10]}")
            excel_data[sub] = {"error": f"column '{col_name}' not found. Available: {avail[:8]}"}
            wb.close()
            continue

        print(f"  Column '{col_name}' found at index {col_idx}")

        # Extract company names from data rows
        data_rows = all_rows[hdr_row_idx + 1:]
        total_rows = len(data_rows)
        names_raw = []
        for row in data_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip()
                if val and val.lower() not in ("none", "n/a", "", "empresa"):
                    names_raw.append(val)

        non_empty = len(names_raw)
        names_norm_set = set(norm_name(n) for n in names_raw if norm_name(n))
        unique = len(names_norm_set)

        print(f"  Data rows: {total_rows} | Non-empty: {non_empty} | Unique (norm): {unique}")
        print(f"  Sample names: {names_raw[:4]}")

        excel_data[sub] = {
            "total_rows": total_rows,
            "non_empty": non_empty,
            "unique_norm": unique,
            "names_norm_set": names_norm_set,
            "names_raw_sample": names_raw[:5],
        }
        wb.close()

    except Exception as e:
        print(f"  ERROR: {e}")
        excel_data[sub] = {"error": str(e)}


# ─── STEP 2: SUPABASE CURRENT STATE ──────────────────────────────────────────

print("\n\n" + "=" * 70)
print("STEP 2: QUERYING SUPABASE")
print("=" * 70)

# 2a. Get all sub_lineas
print("\n[2a] Fetching sub_lineas_negocio...")
sl_data, sl_total, sl_err = supabase_get("sub_lineas_negocio", {"select": "id,codigo,nombre"})
if sl_err:
    print(f"  ERROR: {sl_err}")
    sub_lineas_map = {}
else:
    print(f"  Found {len(sl_data)} sub_lineas:")
    sub_lineas_map = {}
    for sl in sl_data:
        sub_lineas_map[sl["codigo"]] = {"id": sl["id"], "nombre": sl["nombre"]}
        print(f"    id={sl['id']}  codigo={sl['codigo']}  nombre={sl['nombre']}")

# 2b. Count and fetch company names per sub_linea
print("\n[2b] Companies per sub_linea...")
db_data = {}

for sl_code, sl_info in sub_lineas_map.items():
    sl_id = sl_info["id"]
    print(f"\n  [{sl_code}] id={sl_id}")

    esl_data, esl_total, esl_err = supabase_get(
        "empresa_sub_lineas",
        {
            "select": "empresa_id,empresas(id,company_name,company_name_norm)",
            "sub_linea_id": f"eq.{sl_id}",
            "limit": "2000",
        },
    )

    if esl_err:
        print(f"    ERROR (nested join): {esl_err}")
        # Fallback: count only
        esl_data2, esl_total2, esl_err2 = supabase_get(
            "empresa_sub_lineas",
            {"select": "empresa_id", "sub_linea_id": f"eq.{sl_id}", "limit": "2000"},
        )
        if esl_err2:
            print(f"    ERROR (simple): {esl_err2}")
            db_data[sl_code] = {"error": esl_err}
        else:
            count = esl_total2 if esl_total2 is not None else len(esl_data2 or [])
            print(f"    Count (no names): {count}")
            db_data[sl_code] = {"count": count, "names_norm_set": set(), "no_names": True}
        continue

    raw_count = esl_total if esl_total is not None else len(esl_data or [])

    names_norm_set = set()
    names_raw_sample = []
    if esl_data:
        for row in esl_data:
            empresa = row.get("empresas")
            if empresa:
                n = empresa.get("company_name_norm") or empresa.get("company_name") or ""
                normalized = norm_name(n)
                if normalized:
                    names_norm_set.add(normalized)
                if len(names_raw_sample) < 5:
                    names_raw_sample.append(empresa.get("company_name", ""))

    print(f"    Rows: {len(esl_data)} | Content-Range total: {raw_count} | Unique norm names: {len(names_norm_set)}")
    if names_raw_sample:
        print(f"    Sample: {names_raw_sample[:3]}")

    db_data[sl_code] = {
        "count": len(names_norm_set),
        "raw_count": raw_count,
        "names_norm_set": names_norm_set,
    }

# 2c. Total distinct companies
print("\n[2c] Total distinct companies in empresas table...")
emp_data, emp_total, emp_err = supabase_get("empresas", {"select": "id", "limit": "1"})
if emp_err:
    print(f"  ERROR: {emp_err}")
    total_empresas = "ERROR"
else:
    total_empresas = emp_total
    print(f"  Total (Content-Range): {total_empresas}")


# ─── STEP 3: GAP ANALYSIS ────────────────────────────────────────────────────

print("\n\n" + "=" * 70)
print("STEP 3: GAP ANALYSIS")
print("=" * 70)

gap_results = {}
all_sub_lineas_sorted = [cfg["sub_linea"] for cfg in EXCEL_FILES]

for sl_code in all_sub_lineas_sorted:
    excel_info = excel_data.get(sl_code, {})
    db_info = db_data.get(sl_code, {})

    if sl_code not in sub_lineas_map:
        gap_results[sl_code] = {"db_status": "sub_linea NOT in DB"}
        continue

    if "error" in excel_info:
        gap_results[sl_code] = {
            "db_status": "ok",
            "excel_error": excel_info["error"],
            "db_count": db_info.get("count", 0),
        }
        continue

    excel_set = excel_info.get("names_norm_set", set())
    db_set = db_info.get("names_norm_set", set())

    missing = sorted(excel_set - db_set)
    extra = sorted(db_set - excel_set)

    gap_results[sl_code] = {
        "db_status": "ok",
        "excel_count": excel_info.get("unique_norm", 0),
        "db_count": db_info.get("count", 0),
        "missing_count": len(missing),
        "extra_count": len(extra),
        "missing": missing,
        "extra": extra,
    }


# ─── STEP 4: CROSS-SUBLÍNEA DUPLICATES IN EXCEL ──────────────────────────────

print("[4] Cross-sub-línea duplicates in Excel...")
name_to_subs = defaultdict(list)
for sl_code, info in excel_data.items():
    if "error" not in info:
        for n in info.get("names_norm_set", set()):
            name_to_subs[n].append(sl_code)

cross_dupes = {n: sorted(subs) for n, subs in name_to_subs.items() if len(subs) > 1}
print(f"  Companies in 2+ Excel files: {len(cross_dupes)}")


# ─── FINAL REPORT ────────────────────────────────────────────────────────────

SEP = "=" * 70
DASH = "-" * 70

print(f"\n\n{SEP}")
print("=== EXCEL COUNTS ===")
print(SEP)
print(f"  {'Sub-línea':<24} {'Excel rows':>10} {'Non-empty':>10} {'Unique names':>13}")
print(f"  {'-'*24} {'-'*10} {'-'*10} {'-'*13}")
total_excel = 0
for cfg in EXCEL_FILES:
    sl = cfg["sub_linea"]
    info = excel_data.get(sl, {})
    if "error" in info:
        print(f"  {sl:<24} {'ERROR':>10}  {info['error'][:50]}")
    else:
        u = info.get("unique_norm", 0)
        total_excel += u
        print(
            f"  {sl:<24} {info.get('total_rows',0):>10,} {info.get('non_empty',0):>10,} {u:>13,}"
        )
print(f"\n  Total unique company names across all Excels (before dedup): {total_excel:,}")

print(f"\n{SEP}")
print("=== SUPABASE CURRENT STATE ===")
print(SEP)
print(f"  {'Sub-línea':<24} {'Companies in DB':>16}  Status")
print(f"  {'-'*24} {'-'*16}  {'-'*20}")
total_db = sum(v.get("count", 0) for v in db_data.values())
for cfg in EXCEL_FILES:
    sl = cfg["sub_linea"]
    if sl not in sub_lineas_map:
        print(f"  {sl:<24} {'—':>16}  sub_linea NOT CREATED in DB")
    else:
        db_info = db_data.get(sl, {})
        if "error" in db_info and "count" not in db_info:
            print(f"  {sl:<24} {'ERROR':>16}  {db_info.get('error','')[:40]}")
        else:
            cnt = db_info.get("count", 0)
            raw = db_info.get("raw_count", cnt)
            note = f" (pivot rows: {raw})" if raw != cnt else ""
            print(f"  {sl:<24} {cnt:>16,}{note}")

print(f"\n  Total distinct companies in empresas table: {total_empresas}")
print(f"  Sum of per-sub-línea DB counts (with overlap): {total_db:,}")

print(f"\n{SEP}")
print("=== GAP ANALYSIS ===")
print(SEP)
print(f"  {'Sub-línea':<24} {'In Excel':>9} {'In DB':>7} {'Missing':>8} {'Extra in DB':>12}")
print(f"  {'-'*24} {'-'*9} {'-'*7} {'-'*8} {'-'*12}")
total_missing = 0
for cfg in EXCEL_FILES:
    sl = cfg["sub_linea"]
    g = gap_results.get(sl, {})
    if g.get("db_status") == "sub_linea NOT in DB":
        xc = excel_data.get(sl, {}).get("unique_norm", "?")
        print(f"  {sl:<24} {str(xc):>9} {'—':>7}  sub_linea NOT IN DB")
    elif "excel_error" in g:
        print(f"  {sl:<24} {'ERR':>9} {g.get('db_count',0):>7}  Excel error")
    else:
        m = g.get("missing_count", 0)
        total_missing += m
        print(
            f"  {sl:<24} {g.get('excel_count',0):>9,} {g.get('db_count',0):>7,}"
            f" {m:>8,} {g.get('extra_count',0):>12,}"
        )

print(f"\n  Total missing companies (excl. logistica): {total_missing:,}")

print(f"\n{SEP}")
print("=== COMPANIES MISSING PER SUB-LÍNEA (first 20 each) ===")
print(SEP)
for cfg in EXCEL_FILES:
    sl = cfg["sub_linea"]
    g = gap_results.get(sl, {})

    if g.get("db_status") == "sub_linea NOT in DB":
        xl_names = sorted(excel_data.get(sl, {}).get("names_norm_set", set()))
        print(f"\n[{sl}] — sub_linea NOT in DB — {len(xl_names)} companies in Excel:")
        for n in xl_names[:20]:
            print(f"    {n}")
        if len(xl_names) > 20:
            print(f"    ... and {len(xl_names) - 20} more")
    elif "excel_error" in g:
        print(f"\n[{sl}] — Excel error: {g['excel_error']}")
    else:
        missing = g.get("missing", [])
        if missing:
            print(f"\n[{sl}] — {len(missing)} missing from DB:")
            for n in missing[:20]:
                print(f"    {n}")
            if len(missing) > 20:
                print(f"    ... and {len(missing) - 20} more")
        else:
            print(f"\n[{sl}] — 0 missing (all Excel companies present in DB)")

print(f"\n{SEP}")
print("=== EXTRA IN DB (not in any Excel for this sub-línea) ===")
print(SEP)
any_extra = False
for cfg in EXCEL_FILES:
    sl = cfg["sub_linea"]
    g = gap_results.get(sl, {})
    extra = g.get("extra", [])
    if extra:
        any_extra = True
        print(f"\n[{sl}] — {len(extra)} extra (in DB, not in Excel):")
        for n in extra[:20]:
            print(f"    {n}")
        if len(extra) > 20:
            print(f"    ... and {len(extra) - 20} more")
if not any_extra:
    print("  No extras found.")

print(f"\n{SEP}")
print("=== CROSS-SUBLÍNEA DUPLICATES IN EXCEL ===")
print(SEP)
print(f"Companies appearing in 2+ Excel files: {len(cross_dupes)}")
if cross_dupes:
    print()
    sorted_dupes = sorted(cross_dupes.items(), key=lambda x: (-len(x[1]), x[0]))
    for name, subs in sorted_dupes[:60]:
        print(f"  {name:<45} in: {', '.join(subs)}")
    if len(sorted_dupes) > 60:
        print(f"  ... and {len(sorted_dupes) - 60} more")

print(f"\n{SEP}")
print("AUDIT COMPLETE")
print(SEP)
