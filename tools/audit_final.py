"""
Matec Radar B2B — Complete Audit (Read-only)
Excel Prospection Files vs Supabase Production
"""
import sys
import io
import unicodedata
import openpyxl
import requests
from pathlib import Path
from collections import defaultdict

# UTF-8 safe output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE_PATH = Path("c:/Users/Juan/Documents/Agentic Workflows/clients/docs/PROSPECCIÓN")
SUPABASE_URL = "https://supabase.valparaiso.cafe"
SERVICE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJyb2xlIjoic2VydmljZV9yb2xlIiwiaXNzIjoic3VwYWJhc2UiLCJpYXQiOjE3MDAwMDAwMDAsImV4cCI6MjIwMDAwMDAwMH0"
    ".vBDC2y9ofT_-tvJxNdtRtFilgmkFN-ktiSl2AHiZZ3o"
)
HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Accept-Profile": "matec_radar",
    "Prefer": "count=exact",
}

# (sub_linea, file_path, sheet_name, header_row_idx, company_col_name)
# Notes on structures confirmed by inspection:
# - aeropuertos:          header row 0, col "COMPANY NAME"
# - cargo_uld:            header row 1, col "Empresa" (one row per empresa*pais)
# - carton_corrugado:     header row 3, col "COMPANY NAME" (rows 0-2 = merged labels)
# - final_linea:          header row 4, col "COMPANY NAME" (rows 0-3 = merged labels)
# - ensambladoras_motos:  header row 0, col "COMPANY NAME"
# - solumat:              header row 3, col "COMPANY NAME"
# - logistica:            header row 3, col "COMPANY NAME"
CONFIGS = [
    (
        "aeropuertos",
        BASE_PATH / "Linea Aeropuertos" / "BASE DE DATOS AEROPUERTOS FINAL.xlsx",
        "Base de Datos", 0, "COMPANY NAME",
    ),
    (
        "cargo_uld",
        BASE_PATH / "Línea Cargo" / "BASE DE DATOS CARGO LATAM.xlsx",
        "CRM_Cargo_ULD_LATAM", 1, "Empresa",
    ),
    (
        "carton_corrugado",
        BASE_PATH / "Linea Carton y Papel" / "BASE DE DATOS CARTON Y PAPEL.xlsx",
        "Base de Datos", 3, "COMPANY NAME",
    ),
    (
        "final_linea",
        BASE_PATH / "Final de Línea" / "BASE DE DATOS FINAL DE LINEA.xlsx",
        "Base de Datos", 4, "COMPANY NAME",
    ),
    (
        "ensambladoras_motos",
        BASE_PATH / "Ensambladora de Motos" / "BASE DE DATOS ENSAMBLADORAS MOTOS LATAM.xlsx",
        "Base de Datos", 0, "COMPANY NAME",
    ),
    (
        "solumat",
        BASE_PATH / "Línea Solumat" / "BASE DE DATOS SOLUMAT.xlsx",
        "Base de Datos", 3, "COMPANY NAME",
    ),
    (
        "logistica",
        BASE_PATH / "Línea Logistica" / "BASE DE DATOS LOGÍSTICA 2026.xlsx",
        "Base de Datos", 3, "COMPANY NAME",
    ),
]


def norm(s: str) -> str:
    if not s:
        return ""
    nfd = unicodedata.normalize("NFD", str(s).strip())
    return nfd.encode("ascii", "ignore").decode("ascii").lower().strip()


def load_excel(path, sheet, hrow, col_name):
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = rows[hrow]
    col_idx = next(
        (i for i, v in enumerate(hdr) if v and str(v).strip().lower() == col_name.lower()),
        None,
    )
    if col_idx is None:
        avail = [str(v).strip() for v in hdr if v]
        return None, [], 0, 0, f"col '{col_name}' not found. Available: {avail[:6]}"
    data = rows[hrow + 1:]
    raw = [
        str(r[col_idx]).strip()
        for r in data
        if col_idx < len(r) and r[col_idx] and str(r[col_idx]).strip().lower() not in ("none", "n/a", "")
    ]
    ns = set(norm(n) for n in raw if norm(n))
    return ns, raw[:5], len(data), len(raw), None


def fetch_db_names(sl_id: int):
    resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/empresa_sub_lineas",
        headers=HEADERS,
        params={
            "select": "empresa_id,empresas(company_name,company_name_norm)",
            "sub_linea_id": f"eq.{sl_id}",
            "limit": "2000",
        },
        timeout=30,
    )
    if resp.status_code != 200:
        return set(), f"HTTP {resp.status_code}"
    ns = set()
    for row in resp.json():
        e = row.get("empresas")
        if e:
            n = e.get("company_name_norm") or e.get("company_name") or ""
            normalized = norm(n)
            if normalized:
                ns.add(normalized)
    return ns, None


# ── STEP 1: EXCEL ─────────────────────────────────────────────────────────────
print("=" * 70)
print("STEP 1: EXCEL FILES")
print("=" * 70)

excel = {}
for sl, path, sheet, hrow, col in CONFIGS:
    print(f"  [{sl}] {path.name}")
    ns, sample, total_rows, non_empty, err = load_excel(path, sheet, hrow, col)
    if err:
        print(f"    ERROR: {err}")
        excel[sl] = {"error": err, "set": set(), "unique": 0, "total_rows": 0, "non_empty": 0}
    else:
        print(f"    Rows={total_rows}, NonEmpty={non_empty}, Unique={len(ns)}, Sample={sample[:3]}")
        excel[sl] = {"set": ns, "unique": len(ns), "total_rows": total_rows, "non_empty": non_empty}

# ── STEP 2: SUPABASE ──────────────────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 2: SUPABASE")
print("=" * 70)

sl_resp = requests.get(
    f"{SUPABASE_URL}/rest/v1/sub_lineas_negocio",
    headers=HEADERS,
    params={"select": "id,codigo,nombre"},
    timeout=30,
)
sl_map = {}
if sl_resp.status_code == 200:
    for row in sl_resp.json():
        sl_map[row["codigo"]] = {"id": row["id"], "nombre": row["nombre"]}
    print(f"  sub_lineas in DB: {list(sl_map.keys())}")
else:
    print(f"  ERROR fetching sub_lineas: {sl_resp.status_code}")

db = {}
for sl, *_ in CONFIGS:
    if sl not in sl_map:
        db[sl] = {"missing_sl": True, "set": set(), "count": 0}
        print(f"  [{sl}] NOT IN DB")
    else:
        ns, err = fetch_db_names(sl_map[sl]["id"])
        if err:
            db[sl] = {"error": err, "set": set(), "count": 0}
            print(f"  [{sl}] ERROR: {err}")
        else:
            db[sl] = {"set": ns, "count": len(ns)}
            print(f"  [{sl}] DB companies: {len(ns)}")

emp_resp = requests.get(
    f"{SUPABASE_URL}/rest/v1/empresas",
    headers=HEADERS,
    params={"select": "id", "limit": "1"},
    timeout=30,
)
total_empresas = "?"
if emp_resp.status_code in (200, 206):
    cr = emp_resp.headers.get("Content-Range", "")
    if "/" in cr:
        total_empresas = cr.split("/")[1]

# ── REPORT ────────────────────────────────────────────────────────────────────
SEP = "=" * 70

print(f"\n\n{SEP}")
print("=== EXCEL COUNTS ===")
print(SEP)
print(f"  {'Sub-linea':<24} {'Excel rows':>10} {'Non-empty':>10} {'Unique':>8}")
print(f"  {'-'*24} {'-'*10} {'-'*10} {'-'*8}")
for sl, *_ in CONFIGS:
    e = excel[sl]
    if e.get("error"):
        print(f"  {sl:<24} ERROR: {e['error'][:40]}")
    else:
        print(f"  {sl:<24} {e['total_rows']:>10,} {e['non_empty']:>10,} {e['unique']:>8,}")

print(f"\n{SEP}")
print("=== SUPABASE CURRENT STATE ===")
print(SEP)
print(f"  {'Sub-linea':<24} {'In DB':>8}  Notes")
print(f"  {'-'*24} {'-'*8}  {'-'*30}")
for sl, *_ in CONFIGS:
    d = db[sl]
    if d.get("missing_sl"):
        print(f"  {sl:<24} {'—':>8}  sub_linea NOT IN DB (migration pending)")
    elif d.get("error"):
        print(f"  {sl:<24} {'ERR':>8}  {d['error']}")
    else:
        print(f"  {sl:<24} {d['count']:>8,}")
print(f"\n  Total distinct companies in empresas table: {total_empresas}")

print(f"\n{SEP}")
print("=== GAP ANALYSIS ===")
print(SEP)
print(f"  {'Sub-linea':<24} {'In Excel':>8} {'In DB':>6} {'Missing':>8} {'Extra in DB':>12}")
print(f"  {'-'*24} {'-'*8} {'-'*6} {'-'*8} {'-'*12}")
total_missing = 0
for sl, *_ in CONFIGS:
    e = excel[sl]
    d = db[sl]
    if d.get("missing_sl"):
        print(f"  {sl:<24} {e['unique']:>8,} {'—':>6}  sub_linea NOT IN DB")
    elif e.get("error") or d.get("error"):
        print(f"  {sl:<24} {'ERR':>8} {'ERR':>6}")
    else:
        miss = sorted(e["set"] - d["set"])
        extra = sorted(d["set"] - e["set"])
        total_missing += len(miss)
        print(
            f"  {sl:<24} {e['unique']:>8,} {d['count']:>6,} {len(miss):>8,} {len(extra):>12,}"
        )
print(f"\n  Total companies missing from DB (excl. logistica): {total_missing:,}")

print(f"\n{SEP}")
print("=== COMPANIES MISSING FROM DB PER SUB-LINEA ===")
print(SEP)
for sl, *_ in CONFIGS:
    e = excel[sl]
    d = db[sl]
    if d.get("missing_sl"):
        xl_names = sorted(e["set"])
        print(f"\n[{sl}]  sub_linea NOT IN DB  — {len(xl_names)} companies in Excel (first 25):")
        for n in xl_names[:25]:
            print(f"    {n}")
        if len(xl_names) > 25:
            print(f"    ... and {len(xl_names) - 25} more")
    elif e.get("error"):
        print(f"\n[{sl}]  Excel read error: {e['error']}")
    else:
        miss = sorted(e["set"] - d["set"])
        if not miss:
            print(f"\n[{sl}]  0 missing — all Excel companies found in DB")
        else:
            print(f"\n[{sl}]  {len(miss)} missing from DB:")
            for n in miss[:30]:
                print(f"    {n}")
            if len(miss) > 30:
                print(f"    ... and {len(miss) - 30} more")

print(f"\n{SEP}")
print("=== EXTRA IN DB (present in DB but NOT in Excel) PER SUB-LINEA ===")
print(SEP)
for sl, *_ in CONFIGS:
    e = excel[sl]
    d = db[sl]
    if not d.get("missing_sl") and not e.get("error") and not d.get("error"):
        extra = sorted(d["set"] - e["set"])
        if extra:
            print(f"\n[{sl}]  {len(extra)} extra in DB:")
            for n in extra[:25]:
                print(f"    {n}")
            if len(extra) > 25:
                print(f"    ... and {len(extra) - 25} more")

print(f"\n{SEP}")
print("=== CROSS-SUBLINEA DUPLICATES IN EXCEL ===")
print(SEP)
name_to_sls = defaultdict(list)
for sl, *_ in CONFIGS:
    if not excel[sl].get("error"):
        for n in excel[sl]["set"]:
            name_to_sls[n].append(sl)
dupes = {n: sorted(sls) for n, sls in name_to_sls.items() if len(sls) > 1}
print(f"Companies appearing in 2+ Excel files: {len(dupes)}\n")
for n, sls in sorted(dupes.items(), key=lambda x: (-len(x[1]), x[0])):
    print(f"  {n:<45}  in: {', '.join(sls)}")

print(f"\n{SEP}")
print("AUDIT COMPLETE")
print(SEP)
